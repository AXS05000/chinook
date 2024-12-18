import openai
from django.conf import settings
import openpyxl
from openpyxl.utils import get_column_letter
import os
import re
from django.db.models import Q
from django.db import models
from collections import Counter

import openai
from ia.api_key_loader import get_api_key

# Obter a chave da API do banco de dados
openai_api_key = get_api_key("OpenAI")

openai.api_key = openai_api_key


## model="gpt-4o" - Modelo mais rapido e inteligente - 30 000 TPM
## model="gpt-4o-2024-08-06" - 2 Modelo mais rapido e inteligente - 30 000 TPM
## model="gpt-4" - 2 Modelo mais rapido e inteligente - 10 000 TPM
## model="gpt-3.5-turbo" - Modelo menos inteligente - 60 000 TPM


def get_chat_response(prompt, context=""):
    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    # Enviar uma mensagem clara e estruturada para a API
    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é um assistente útil que auxilia na análise de avaliações de pais sobre uma escola.",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Pergunta do usuário:\n{prompt}"},
        ],
        max_tokens=2050,
    )
    print(f"Total tokens usados: {response['usage']['total_tokens']}")
    formatted_response = response["choices"][0]["message"]["content"].strip()

    # Formatações adicionais
    formatted_response = re.sub(r"###", "<br>", formatted_response)
    formatted_response = re.sub(
        r"\*\*(.*?)\*\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )
    formatted_response = re.sub(
        r"\*(.*?)\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )

    return formatted_response


def calcular_nps(informacoes):
    nps_responses = informacoes.filter(
        questao="Em uma escala de 0 a 10, o quanto você recomendaria a escola para um amigo ou familiar?"
    )
    total_respostas = nps_responses.count()
    if total_respostas == 0:
        return "Não há respostas suficientes para calcular o NPS."

    promotores = nps_responses.filter(resposta__gte=9).count()
    neutros = nps_responses.filter(resposta__in=[7, 8]).count()
    detratores = nps_responses.filter(resposta__lte=6).count()

    nps = ((promotores - detratores) / total_respostas) * 100

    return nps


def obter_distribuicao_nps(informacoes):
    nps_responses = informacoes.filter(
        questao="Em uma escala de 0 a 10, o quanto você recomendaria a escola para um amigo ou familiar?"
    )
    promotores = nps_responses.filter(resposta__gte=9).count()
    neutros = nps_responses.filter(resposta__in=[7, 8]).count()
    detratores = nps_responses.filter(resposta__lte=6).count()

    return promotores, neutros, detratores


def calculate_nps(
    informacoes, suposicao_promotores=0, suposicao_neutros=0, suposicao_detratores=0
):
    total_respostas = (
        informacoes.count()
        + suposicao_promotores
        + suposicao_neutros
        + suposicao_detratores
    )
    if total_respostas == 0:
        return None  # Evita divisão por zero se não houver respostas

    promotores = informacoes.filter(resposta__in=[9, 10]).count() + suposicao_promotores
    detratores = informacoes.filter(resposta__lte=6).count() + suposicao_detratores

    percentual_promotores = (promotores / total_respostas) * 100
    percentual_detratores = (detratores / total_respostas) * 100

    nps = percentual_promotores - percentual_detratores
    return nps


def generate_excel_report(informacoes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Avaliações"

    headers = [
        "Nome",
        "Persona",
        "Data da Resposta",
        "Unidade",
        "Questão",
        "Resposta",
        "Comentário",
    ]
    ws.append(headers)

    for info in informacoes:
        row = [
            info.nome,
            info.persona,
            info.data_resposta,
            info.unidade,
            info.questao,
            info.resposta,
            info.comentario,
        ]
        ws.append(row)

    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    file_name = "relatorio_avaliacoes.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)
    wb.save(file_path)

    return settings.MEDIA_URL + file_name


def respostas_por_dia(informacoes):
    datas = [info.data_resposta for info in informacoes]
    contagem = Counter(datas)
    return contagem


def resumo_por_pergunta(informacoes):
    perguntas = informacoes.values_list("questao", flat=True).distinct()
    resumo = {}

    for pergunta in perguntas:
        respostas = informacoes.filter(questao=pergunta)
        total = respostas.count()
        negativo = respostas.filter(resposta__in=[1, 2]).count()
        neutro = respostas.filter(resposta=3).count()
        positivo = respostas.filter(resposta__in=[4, 5]).count()
        comentarios = respostas.values_list("comentario", flat=True)

        resumo[pergunta] = {
            "total": total,
            "negativo": negativo,
            "neutro": neutro,
            "positivo": positivo,
            "comentarios": list(comentarios),
        }
    return resumo


##############################################################################################



def classify_question(prompt, api_key):
    openai.api_key = api_key

    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[
            {
                "role": "system",
                "content": "Você é um assistente útil que classifica perguntas de funcionários em categorias: 'benefício', 'folha de ponto', 'salário', 'férias' ou 'banco de horas'. Responda apenas com a categoria apropriada.",
            },
            {"role": "user", "content": prompt},
        ],
        max_tokens=20,
    )
    category = response["choices"][0]["message"]["content"].strip().lower()
    # Ajustar para garantir que 'banco de horas' seja tratado como 'folha de ponto'
    if category == "banco de horas":
        category = "folha de ponto"
    return category

def config_chat_rh(prompt, api_key, context=""):
    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    openai.api_key = api_key

    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[
            {
                "role": "system",
                "content": "Você é o assistente Chinook de recursos humanos da Empresa Maple Bear auxiliando o setor Gente Gestão, que auxilia na resposta de perguntas dos funcionários. Observação importante: sempre que for realizar listagem ou fazer uma lista onde tem indicativos de números antes dos números colocar esses 3 símbolos ### e formate todos os links utilizando Markdown da seguinte forma: [texto do link](URL).",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Pergunta do usuário:\n{prompt}"},
        ],
        max_tokens=2050,
    )
    print(f"Total tokens usados: {response['usage']['total_tokens']}")
    formatted_response = response["choices"][0]["message"]["content"].strip()

    # Formatações adicionais
    formatted_response = re.sub(r"###", "<br>", formatted_response)
    formatted_response = re.sub(
        r"\*\*(.*?)\*\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )
    formatted_response = re.sub(
        r"\*(.*?)\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )
    formatted_response = re.sub(
        r"\[(.*?)\]\((.*?)\)",
        r'<a href="\2" target="_blank">\1</a>',
        formatted_response,
    )
    formatted_response = re.sub(
        r'(?<!")\b(https?://[^\s]+)\b(?!")',
        r'<a href="\1" target="_blank">\1</a>',
        formatted_response,
    )

    return formatted_response




################################ CHAT SAF########################################

def classify_question_chat_central(prompt, api_key):
    openai.api_key = api_key
    valid_categories = [
        "informações gerais",
        "pedido",
        "planificador",
        "sprinklr",
        "ouvidoria",
        "nps",
        "cliente_oculto",
        "vendas",
        "base de conhecimento",
        "analise completa da escola"
    ]

    for attempt in range(10):
        response = openai.ChatCompletion.create(
            model="gpt-4o-2024-08-06",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Você é um assistente útil que classifica perguntas sobre escolas em categorias: "
                        "('informações gerais' - Se o usuário pedir informações como resumo da escola, informações básicas como nome, CNPJ, cluster, endereço, avaliações. "
                        "Exemplos de perguntas para essa categoria: Faça um resumo dessa escola. Mas se ele falar 'faça um resumo do cliente oculto ou NPS dessa escola' ou falar 'faça um resumo bem resumido do cliente oculto ou NPS dessa escola' é outra categoria não é essa), "
                        "('pedido' -  Se o usuário pedir consulte esse pedido ou olha esse pedido e o pedido for 10 caracteres e começar com '2' escolha essa categoria.), "
                        "('planificador' -  Se o usuário pedir informações relacionadas ao planificador escolha essa categoria.), "
                        "('sprinklr' -  Se o usuário pedir informações relacionadas a ticket ou à Sprinklr escolha essa categoria, observação pode ter erros de digitação então se ele colocar Sprinklr, Sbrinklr ou outras variações escolha essa categoria.), "
                        "('ouvidoria' - Se o usuário pedir informações relacionadas a ouvidoria ou SAC escolha essa categoria), "
                        "('nps' - Se o usuário pedir informações relacionadas ao Net Promoter Score(NPS) responda com essa categoria), "
                        "('cliente_oculto' - Se o usuário pedir informações relacionadas ao Cliente Oculto responda com essa categoria), "
                        "('vendas' ou 'relatório de vendas' - Se o usuário pedir informações sobre vendas da escola ou relatório de vendas da escola escolha essa categoria.), "
                        "('base de conhecimento' - Se o usuário pedir algo que pareça estar em uma base de conhecimento escolha essa categoria. Por exemplo se na pergunta tiver algo como conforme base de conhecimento, como na base de conhecimento, no conhecimento.), "
                        "('analise completa da escola' - Só escolha essa categoria se tiver esse conjunto de palavras ou algo muito similar. Por exemplo: 'Faça uma análise completa dessa escola', 'Faça uma avaliação geral dessa escola' ou até algo com 'Olhe todas as informações dessa escola'). "
                        "Responda apenas com a categoria apropriada. Se você não conseguir categorizar a pergunta, responda com 'base de conhecimento'."
                    ),
                },
                {"role": "user", "content": prompt},
            ],
            max_tokens=2500,
        )

        print(f"Tentativa {attempt + 1}: Total tokens usados para classificar a categoria: {response['usage']['total_tokens']}")

        category = response["choices"][0]["message"]["content"].strip().lower()

        if category in valid_categories:
            print(f"Categoria válida encontrada: {category}")
            return category

        print(f"Categoria inválida: {category}. Repetindo classificação...")

    raise ValueError("A classificação falhou após 10 tentativas.")

def config_chat_central(prompt, api_key, context=""):
    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    openai.api_key = api_key

    response = openai.ChatCompletion.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": "Você é o assistente Chinook da Empresa Maple Bear auxiliando nas informações das escolas e na resposta de perguntas dos funcionários. Observação importante: sempre que for realizar listagem, um resumo, uma tabela ou fazer uma lista, coloque esses 3 símbolos ### antes de cada tópico e formate todos os links utilizando Markdown da seguinte forma: [texto do link](URL).",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Pergunta do usuário:\n{prompt}"},
        ],
        max_tokens=2050,
    )

    # Extrair os tokens usados da resposta da API
    tokens_used = response['usage']['total_tokens']

    # Formatar a resposta
    formatted_response = response["choices"][0]["message"]["content"].strip()
    formatted_response = re.sub(r"###", "<br>", formatted_response)
    formatted_response = re.sub(r"####", "<br>", formatted_response)
    formatted_response = re.sub(
        r"\*\*(.*?)\*\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )
    formatted_response = re.sub(
        r"\*(.*?)\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )
    formatted_response = re.sub(
        r"\[(.*?)\]\((.*?)\)",
        r'<a href="\2" target="_blank">\1</a>',
        formatted_response,
    )
    formatted_response = re.sub(
        r'(?<!")\b(https?://[^\s]+)\b(?!")',
        r'<a href="\1" target="_blank">\1</a>',
        formatted_response,
    )

    # Retorne tanto o texto formatado quanto o total de tokens usados como um dicionário
    return {
        'formatted_response': formatted_response,
        'tokens': tokens_used
    }


def extract_order_number(prompt, api_key):
    openai.api_key = api_key

    for attempt in range(5):  # Tenta no máximo 5 vezes para evitar loops infinitos
        print(f"Tentativa {attempt + 1} de extração do número do pedido")
        response = openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": "Você é um assistente que extrai números de pedidos de uma pergunta. A pergunta sempre terá um número de pedido, como 'Consulta o pedido 1234567890'. Identifique o número do pedido na pergunta do usuário e retorne apenas o número, sem explicações.",
                },
                {"role": "user", "content": prompt},
            ],
            max_tokens=50,
        )

        order_number = response["choices"][0]["message"]["content"].strip()
        print(f"Resposta da IA: {order_number}")

        # Verifica se o número do pedido tem exatamente 10 caracteres
        if len(order_number) == 10 and order_number.isdigit():
            print(f"Número do pedido válido extraído: {order_number}")
            return order_number
        else:
            print(f"Número inválido extraído: {order_number}. Tentando novamente...")


    raise ValueError("Não foi possível extrair um número de pedido válido após 5 tentativas.")

################################ CHAT EXCOM########################################

def classify_question_excom(prompt, api_key):
    openai.api_key = api_key

    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é um assistente útil que classifica perguntas sobre escolas em categorias: ('ouvidoria' - Se o usuário pedir informações relacionado a ouvidoria ou sac escolha essa categoria),('NPS' - Se o usuário pedir informações relacionado ao Net Promoter Score(NPS) responda com essa categoria), ('cliente_oculto' - Se o usuário pedir informações relacionado ao Cliente Oculto responda com essa categoria). Responda apenas com a categoria apropriada.  Se você não conseguir categorizar a pergunta, responda com 'outros'.",
            },
            {"role": "user", "content": prompt},
        ],
        max_tokens=2500,
    )
    print(f"Total tokens usados para classificar a categoria: {response['usage']['total_tokens']}")
    category = response["choices"][0]["message"]["content"].strip().lower()
    return category

def config_chat_excom(prompt, api_key, context=""):
    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    openai.api_key = api_key

    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é o assistente Excom da Empresa Maple Bear, um consultor de franchising especializado em programa de excelência. Observação importante: sempre que for realizar listagem, um resumo, uma tabela ou fazer uma lista, coloque esses 3 símbolos ### antes de cada tópico e formate todos os links utilizando Markdown da seguinte forma: [texto do link](URL).",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Pergunta do usuário:\n{prompt}"},
        ],
        max_tokens=2050,
    )

    # Extrair os tokens usados da resposta da API
    tokens_used = response['usage']['total_tokens']

    # Formatar a resposta
    formatted_response = response["choices"][0]["message"]["content"].strip()
    formatted_response = re.sub(r"###", "<br>", formatted_response)
    formatted_response = re.sub(r"####", "<br>", formatted_response)
    formatted_response = re.sub(
        r"\*\*(.*?)\*\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )
    formatted_response = re.sub(
        r"\*(.*?)\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )
    formatted_response = re.sub(
        r"\[(.*?)\]\((.*?)\)",
        r'<a href="\2" target="_blank">\1</a>',
        formatted_response,
    )
    formatted_response = re.sub(
        r'(?<!")\b(https?://[^\s]+)\b(?!")',
        r'<a href="\1" target="_blank">\1</a>',
        formatted_response,
    )

    # Retorne tanto o texto formatado quanto o total de tokens usados como um dicionário
    return {
        'formatted_response': formatted_response,
        'tokens': tokens_used
    }



############################################# Resumo NPS###########################################################

def config_resumo_nps(prompt, api_key, context=""):
    print("Iniciando a configuração do resumo NPS...")

    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    openai.api_key = api_key

    print("Enviando a requisição para a API da OpenAI...")
    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é o assistente Chinook da Empresa Maple Bear auxiliando em informações sobre as avaliações feitas pelos pais a escola.",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Pergunta do usuário:\n{prompt}"},
        ],
        max_tokens=2050,
    )
    print(f"Resposta recebida. Total de tokens usados: {response['usage']['total_tokens']}")

    formatted_response = response["choices"][0]["message"]["content"].strip()
    print(f"Resposta formatada: {formatted_response[:50]}...")  # Exibe apenas os primeiros 100 caracteres da resposta

    return formatted_response


### Resumo Geral Positivo e Negativo ###
def config_resumo_nps_geral(prompt, api_key, context=""):
    print("Iniciando a configuração do resumo NPS...")

    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    openai.api_key = api_key

    print("Enviando a requisição para a API da OpenAI...")
    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é o assistente Chinook da Empresa Maple Bear auxiliando em informações sobre as avaliações feitas pelos pais a escola.",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Pergunta do usuário:\n{prompt}"},
        ],
        max_tokens=2050,
    )
    print(f"Resposta recebida. Total de tokens usados: {response['usage']['total_tokens']}")

    formatted_response = response["choices"][0]["message"]["content"].strip()
    print(f"Resposta formatada: {formatted_response[:50]}...")  # Exibe apenas os primeiros 100 caracteres da resposta

    return formatted_response



############################################# Resumo SAC Ouvidoria ###########################################################


def config_resumo_sac(prompt, api_key, context=""):
    print("Iniciando a configuração do resumo SAC...")

    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    openai.api_key = api_key

    print("Enviando a requisição para a API da OpenAI...")
    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é o assistente Chinook da Empresa Maple Bear auxiliando em informações sobre os atendimentos do SAC",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Pergunta do usuário:\n{prompt}"},
        ],
        max_tokens=2050,
    )
    print(f"Resposta recebida. Total de tokens usados: {response['usage']['total_tokens']}")

    formatted_response = response["choices"][0]["message"]["content"].strip()
    print(f"Resposta formatada: {formatted_response[:50]}...")  # Exibe apenas os primeiros 100 caracteres da resposta

    return formatted_response



def processar_resumos_sac(context, api_key):
    print("Iniciando o processamento dos resumos SAC...")

    openai.api_key = api_key

    print("Enviando os resumos para a API da OpenAI...")
    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",  # Certifique-se de que esse modelo seja o correto para o seu uso
        messages=[
            {
                "role": "system",
                "content": "Você é um assistente que auxilia na identificação de escolas relevantes baseadas em resumos de SAC",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": "Por favor, liste os IDs das escolas que parecem ser relevantes com base nos resumos fornecidos."},
        ],
        max_tokens=1500
    )

    print(f"Resposta recebida da API. Total de tokens usados: {response['usage']['total_tokens']}")

    formatted_response = response["choices"][0]["message"]["content"].strip()
    print(f"Resposta da API formatada: {formatted_response[:50]}...")  # Exibindo os primeiros 50 caracteres

    # Supondo que a API retorna algo como "IDs relevantes: 1, 2, 3"
    escolas_relevantes_ids = [int(id.strip()) for id in formatted_response.split(",") if id.strip().isdigit()]

    return escolas_relevantes_ids

def validar_detalhes_ouvidoria(context, api_key):
    print("Iniciando a validação detalhada das escolas relacionadas...")

    openai.api_key = api_key

    print("Enviando os detalhes das escolas relacionadas para a API da OpenAI...")
    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é um assistente que verifica a relevância de informações detalhadas sobre reclamações de SAC de escolas.",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": "Identifique os IDs das escolas que realmente possuem informações relacionadas ao assunto em questão."},
        ],
        max_tokens=1500
    )

    print(f"Resposta recebida da API. Total de tokens usados: {response['usage']['total_tokens']}")

    formatted_response = response["choices"][0]["message"]["content"].strip()
    print(f"Resposta da API formatada: {formatted_response[:50]}...")  # Exibindo os primeiros 50 caracteres

    # Supondo que a API retorna algo como "IDs relevantes: 1, 2, 3"
    escolas_validadas_ids = [int(id.strip()) for id in formatted_response.split(",") if id.strip().isdigit()]

    return escolas_validadas_ids

############################################# Resumo Cliente Oculto###########################################################



def config_resumo_cliente_oculto(prompt, api_key, context=""):
    print("Iniciando a configuração do resumo do Cliente Oculto 2024...")

    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    openai.api_key = api_key

    print("Enviando a requisição para a API da OpenAI...")
    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é o assistente Chinook da Empresa Maple Bear auxiliando em informações sobre as avaliações do Cliente Oculto. Cliente Oculto é uma avaliação realizada por cliente secreto enviado pela franqueada Maple Bear onde o objetivo é avaliar a escola do ponte de vista de um cliente",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Pergunta do usuário:\n{prompt}"},
        ],
        max_tokens=2050,
    )
    print(f"Resposta recebida. Total de tokens usados: {response['usage']['total_tokens']}")

    formatted_response = response["choices"][0]["message"]["content"].strip()
    print(f"Resposta formatada: {formatted_response[:50]}...")  # Exibe apenas os primeiros 50 caracteres da resposta

    return formatted_response


############################################# CONFIG RESUMO VISITA ###########################################################


def config_resumo_visita_escola(prompt, api_key, context=""):
    print("Iniciando a configuração do resumo da Visita Escola...")

    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    openai.api_key = api_key

    print("Enviando a requisição para a API da OpenAI...")
    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é o assistente Chinook da Empresa Maple Bear auxiliando em informações sobre as visitas às escolas. Resuma o comentário da visita de forma clara e objetiva.",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Comentário da visita:\n{prompt}"},
        ],
        max_tokens=2050,
    )
    print(f"Resposta recebida. Total de tokens usados: {response['usage']['total_tokens']}")

    formatted_response = response["choices"][0]["message"]["content"].strip()
    print(f"Resposta formatada: {formatted_response[:50]}...")  # Exibe apenas os primeiros 50 caracteres da resposta

    return formatted_response


############################################# CHAT CENTRAL###########################################################


def config_simple_chat(prompt, context=""):
    if not context:
        context = "No relevant information found in the database."
    if not prompt:
        prompt = "No question provided."

    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é o assistente simples da Empresa Maple Bear auxiliando na informações das escolas e na resposta de perguntas dos funcionários. Observação importante: sempre que for realizar listagem, um resumo, uma tabela ou fazer uma lista colocar esses 3 símbolos ### antes de cada tópico e formate todos os links utilizando Markdown da seguinte forma: [texto do link](URL).",
            },
            {"role": "user", "content": f"Contexto:\n{context}"},
            {"role": "user", "content": f"Pergunta do usuário:\n{prompt}"},
        ],
        max_tokens=2050,
    )
    print(f"Total tokens usados: {response['usage']['total_tokens']}")
    formatted_response = response["choices"][0]["message"]["content"].strip()

    # Formatações adicionais
    formatted_response = re.sub(r"###", "<br>", formatted_response)
    formatted_response = re.sub(r"####", "<br>", formatted_response)

    formatted_response = re.sub(
        r"\*\*(.*?)\*\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )
    formatted_response = re.sub(
        r"\*(.*?)\*",
        r"<span style='font-weight: bold;'>\1</span>",
        formatted_response,
    )
    formatted_response = re.sub(
        r"\[(.*?)\]\((.*?)\)",
        r'<a href="\2" target="_blank">\1</a>',
        formatted_response,
    )
    formatted_response = re.sub(
        r'(?<!")\b(https?://[^\s]+)\b(?!")',
        r'<a href="\1" target="_blank">\1</a>',
        formatted_response,
    )

    return formatted_response



############################################# RESUMO ALTERAÇÕES PLANIFICADOR ###########################################################


def config_resumo_alteracoes(prompt, api_key):
    openai.api_key = api_key

    print("Enviando a requisição para a API da OpenAI...")

    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {
                "role": "system",
                "content": "Você é um assistente que irá resumir todas as alterações feitas, não precisa colocar o que estava antes e sim apenas os campos atuais. Caso tenha essa linha '- Usuário Modificação: None' ou as linhas de data de atualização dos blocos pode desconsiderar ela. Mas coloque um texto como se fosse estive fazendo um report das atividades para o seu gestor. Nos casos das colunas CRM B2C, Circular de Oferta 2025, Toddle e Árvore se estiver SIM quer dizer que foi implementado, então para esses casos coloque 'implementação realizada na escola' apenas no caso da Circular que tem que ser Publicação da Circular realizada. OBS: Toddle e Árvore são no feminino. Nos nomes das escolas não coloque ' - ' deixe apenas o nome. Todas as datas coloque no formato brasileiro dessa forma exemplo: '27/09/2024' e não coloque NUNCA ponto final nas frases, exemplo: Não coloque 'atualizada para 24/09/2024.' coloque 'atualizada para 24/09/2024;'.",
            },
            {"role": "user", "content": f"Resuma as seguintes alterações:\n{prompt}"},
        ],
        max_tokens=1000,
    )

    formatted_response = response["choices"][0]["message"]["content"].strip()
    print(f"Resposta formatada: {formatted_response[:50]}...")  # Exibe os primeiros 50 caracteres

    return formatted_response



############################################# RESUMO BASE DE CONHECIMENTO ###########################################################

def gerar_resumo_base_de_conhecimento(conteudo, api_key):
    """
    Gera um resumo detalhado do conteúdo para garantir que todas as informações importantes
    estejam presentes na coluna resumo.
    """
    print("Iniciando a geração de resumo completo...")
    openai.api_key = api_key

    # Prompt ajustado para garantir que todo o conteúdo seja incluído no resumo
    prompt = (
        "Você é um assistente que deve gerar um resumo detalhado e completo do texto fornecido. "
        "O resumo deve capturar todas as informações presentes no texto original sem omitir nada relevante, "
        "e deve ser conciso, mas suficiente para garantir que todas as ideias principais estejam presentes. "
        "Segue o texto:\n"
        f"{conteudo}\n"
        "Por favor, crie um resumo que inclua todas as informações essenciais do texto."
    )

    # Chamada à API
    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {"role": "system", "content": "Você é um especialista em gerar resumos detalhados e completos."},
            {"role": "user", "content": prompt},
        ],
        max_tokens=2000,  # Ajuste o limite para garantir resumos completos
    )

    # Extrai o resumo gerado
    resumo = response["choices"][0]["message"]["content"].strip()
    print(f"Resumo gerado: {resumo[:100]}...")  # Exibe os primeiros 100 caracteres para validação
    return resumo

def filtrar_resumos_conhecimento(pergunta, context_resumos, api_key):
    """
    Filtra os IDs relevantes com base nos resumos disponíveis e na pergunta do usuário.
    """
    print("Filtrando resumos relevantes para a pergunta...")
    openai.api_key = api_key

    # Contexto detalhado com a pergunta e os resumos
    prompt = (
        "Você é um assistente especializado em encontrar conteúdos relevantes com base em resumos.\n"
        "Abaixo está uma pergunta do usuário e uma lista de resumos identificados por IDs.\n"
        "Retorne os IDs dos resumos que são mais relevantes para responder à pergunta.\n\n"
        f"Pergunta: {pergunta}\n\n"
        "Resumos:\n"
        f"{context_resumos}\n\n"
        "Por favor, liste os IDs relevantes separados por vírgulas, como no formato: 1849, 1854, 1855."
    )

    response = openai.ChatCompletion.create(
        model="gpt-4o-2024-08-06",
        messages=[
            {"role": "system", "content": "Você é um especialista em análise de relevância baseada em resumos."},
            {"role": "user", "content": prompt},
        ],
        max_tokens=500,
    )

    formatted_response = response["choices"][0]["message"]["content"].strip()
    print(f"Resposta da API: {formatted_response}")

    # Extrai apenas números separados por vírgulas
    try:
        ids_relevantes = [
            int(num) for num in formatted_response.replace(" ", "").split(",") if num.isdigit()
        ]
    except Exception as e:
        print(f"Erro ao processar IDs: {str(e)}")
        ids_relevantes = []

    print(f"IDs relevantes extraídos: {ids_relevantes}")
    return ids_relevantes