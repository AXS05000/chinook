import json
import random
import markdown
import time
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
from .forms import PlanificadorForm, AtualizarIDEscolaForm, ReclamacaoForm
from datetime import datetime
from datetime import timedelta
from django.forms.models import model_to_dict
from django.utils import timezone
from django.db.models import Count
from django.db.models import Q 
from usuarios.models import UserRequestLog
from usuarios.models import CustomUsuario
from decimal import Decimal
import locale
from glex.models import Base_de_Conhecimento_Geral
from .models import (
    Informacao,
    RegistroIA,
    Resumo_SAC,
    Ouvidoria_SAC,
    Processo, 
    Protesto,
    Beneficio,
    Ticket_Sprinklr,
    Reclamacao,
    FolhaPonto,
    HistoricoAlteracoes,
    Resumo_Respostas_NPS,
    Pedido,
    Salario,
    Resumo_Respostas_NPS_1_Onda_Geral,
    Ferias,
    Resumo_Visita_Escola,
    ResumoAlteracoes_Planificador,
    Visita_Escola,
    CRM_FUI,
    Respostas_NPS,
    Vendas_SLM_2024,
    Vendas_SLM_2025,
    Base_de_Conhecimento,
    Planificador_2024,
    PedidosAlterados,
    Avaliacao_Cliente_Oculto_24,
    Resumo_Respostas_ClienteOculto24,
)
from django.db import transaction
from django.shortcuts import render, redirect
from django.contrib import messages
import pandas as pd
from bs4 import BeautifulSoup
from .utils import (
    get_chat_response,
    calcular_nps,
    processar_resumos_sac,
    obter_distribuicao_nps,
    classify_question_excom,
    config_chat_rh,
    extract_order_number,
    gerar_resumo_base_de_conhecimento,
    filtrar_resumos_conhecimento,
    config_chat_excom,
    config_resumo_sac,
    config_resumo_visita_escola,
    classify_question,
    config_simple_chat,
    validar_detalhes_ouvidoria,
    config_resumo_cliente_oculto,
    config_resumo_alteracoes,
    config_resumo_nps,
    config_chat_central,
    classify_question_chat_central,
)
import openpyxl
from openpyxl.utils import get_column_letter
from django.views import View
from django.views.generic import ListView


class ExcelImportView(View):
    def get(self, request):
        # Retorna o template de upload de arquivo
        return render(request, "chatapp/importar_nps.html")

    def post(self, request):
        excel_file = request.FILES["excel_file"]

        # Carrega o arquivo Excel na memória
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Itera sobre as linhas do arquivo Excel
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignora o cabeçalho
            nome = row[0]
            persona = row[1]
            data_resposta = row[2]
            unidade = row[3]
            questao = row[4]
            resposta = row[5]
            comentario = row[6]

            if (
                nome
                and persona
                and data_resposta
                and unidade
                and questao
                and resposta is not None
            ):
                # Cria um novo objeto Informacao
                Informacao.objects.create(
                    nome=nome,
                    persona=persona,
                    data_resposta=data_resposta,
                    unidade=unidade,
                    questao=questao,
                    resposta=resposta,
                    comentario=comentario,
                )

        return HttpResponse("Importação realizada com sucesso!")


@csrf_exempt
def chat_view(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            user_message = data.get("message", "").lower()
            context = ""

            # Escalas de avaliação
            escalas = {
                "Metodologia de ensino": "1 - Péssimo, 2 - Ruim, 3 - Mediano, 4 - Bom, 5 - Excelente",
                "Infraestrutura": "1 - Péssimo, 2 - Ruim, 3 - Mediano, 4 - Bom, 5 - Excelente",
                "Atendimento pedagógico": "1 - Péssimo, 2 - Ruim, 3 - Mediano, 4 - Bom, 5 - Excelente",
                "Atendimento administrativo/financeiro": "1 - Péssimo, 2 - Ruim, 3 - Mediano, 4 - Bom, 5 - Excelente",
                "Em uma escala de 0 a 10, o quanto você recomendaria a escola para um amigo ou familiar?": "0 - Horrível a 10 - Perfeito",
            }

            # Consultar o banco de dados para obter informações relevantes
            informacoes = Informacao.objects.all()
            total_respostas = informacoes.count()
            if total_respostas > 0:
                soma_respostas = sum(info.resposta for info in informacoes)
                media_respostas = soma_respostas / total_respostas
                context_list = []
                for info in informacoes:
                    escala = escalas.get(info.questao, "")
                    context_list.append(
                        f"Nome: {info.nome}, Persona: {info.persona}, Data da Resposta: {info.data_resposta}, Unidade: {info.unidade}, Questão: {info.questao} (Escala: {escala}), Resposta: {info.resposta}, Comentário: {info.comentario}"
                    )
                context = "\n".join(context_list)
                context += (
                    f"\n\nA média das respostas para a escola é: {media_respostas:.2f}"
                )
            else:
                context = "No relevant information found in the database."

            # Checar se a pergunta é sobre a distribuição de NPS
            if "quantos" in user_message and (
                "promotores" in user_message
                or "detratores" in user_message
                or "neutros" in user_message
            ):
                promotores, neutros, detratores = obter_distribuicao_nps(informacoes)
                response = f"No NPS, houve {promotores} promotores, {neutros} neutros e {detratores} detratores."
            # Checar se a pergunta é sobre o valor do NPS
            elif any(
                term in user_message
                for term in ["nps", "promotores", "detratores", "passivas", "neutras"]
            ):
                nps_score = calcular_nps(informacoes)
                response = f"O Net Promoter Score (NPS) da escola é: {nps_score:.2f}"
            else:
                # Obter resposta do ChatGPT
                response = get_chat_response(user_message, context)

            return JsonResponse({"response": response})
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
    return render(request, "chatapp/chat.html")




####################################################################################################################

@csrf_exempt
@login_required(login_url='/login/')
def hr_assistant_view(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            user_message = data.get("message", "").lower()
            user = request.user

            if not user.api_key:
                return JsonResponse({"error": "Usuário não possui chave API válida"}, status=403)

            api_key = user.api_key
            context = ""

            # Classificar a pergunta do usuário
            tipo_informacao = classify_question(user_message, api_key)

            # Definir os modelos a serem verificados
            modelos = {
                "benefício": Beneficio,
                "folha de ponto": FolhaPonto,
                "salário": Salario,
                "férias": Ferias,
            }

            # Buscar a informação correta baseada no tipo de informação identificado
            if tipo_informacao in modelos:
                modelo = modelos[tipo_informacao]
                informacoes = modelo.objects.all()
                if informacoes.exists():
                    context = informacoes.first().descricao
                    response = config_chat_rh(user_message, api_key, context)
                else:
                    context = f"No relevant information found for {tipo_informacao} in the database."
                    response = config_chat_rh(user_message, api_key, context)
            else:
                context = f"No relevant information found for {tipo_informacao} in the database."
                response = config_chat_rh(user_message, api_key, context)

            # Se a resposta não for satisfatória ou não houver informações, buscar em todas as categorias
            if "No relevant information found" in context or not context:
                context_list = []
                for key, model in modelos.items():
                    informacoes = model.objects.all()
                    if informacoes.exists():
                        context_list.append(informacoes.first().descricao)
                if context_list:
                    context = "\n".join(context_list)
                    response = config_chat_rh(user_message, api_key, context)

            return JsonResponse({"response": response})
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
    return render(request, "chatapp/hr_chat.html")

################################################# IMPORTAR FUI######################################################

@login_required(login_url='/login/')
def import_crm_fui(request):
    if request.method == "POST":
        file = request.FILES["file"]
        if not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_crm_fui")

        df = pd.read_excel(file)
        for _, row in df.iterrows():
            CRM_FUI.objects.update_or_create(
                id_escola=row["id_escola"],
                defaults={
                    "nome_da_escola": row["nome_da_escola"],
                    "CNPJ": row["CNPJ"],
                    "status_da_escola": row["status_da_escola"],
                    "slms_vendidos": row["slms_vendidos"],
                    "slms_vendidos_25": row["slms_vendidos_25"],
                    "meta": row["meta"],
                    "meta_adiantamento_24_25": row["meta_adiantamento_24_25"],
                    "cluster": row["cluster"],
                    "cep_escola": row["cep_escola"],
                    "endereco": row["endereco"],
                    "complemento_escola": row["complemento_escola"],
                    "bairro_escola": row["bairro_escola"],
                    "cidade_da_escola": row["cidade_da_escola"],
                    "estado_da_escola": row["estado_da_escola"],
                    "regiao_da_escola": row["regiao_da_escola"],
                    "status_de_adimplencia": row["status_de_adimplencia"],
                    "inadimplencia": row["inadimplencia"],
                    "ticket_medio": row["ticket_medio"],
                    "valor_royalties": row["valor_royalties"],
                    "valor_fdmp": row["valor_fdmp"],
                    "segmento_da_escola": row["segmento_da_escola"],
                    "atual_serie": row["atual_serie"],
                    "avanco_segmento": row["avanco_segmento"],
                    "telefone_de_contato_da_escola": row[
                        "telefone_de_contato_da_escola"
                    ],
                    "email_da_escola": row["email_da_escola"],
                    "nps_pais_2024_1_onda": row["nps_pais_2024_1_onda"],
                    "quality_assurance_2024": row["quality_assurance_2024"],
                    "cliente_oculto_2024": row["cliente_oculto_2024"],
                    "consultor_saf": row["consultor_saf"],
                    "consultor_academico": row["consultor_academico"],
                    "consultor_comercial": row["consultor_comercial"],
                    "consultor_gestao_escolar": row["consultor_gestao_escolar"],
                    "dias_uteis_entrega_slm": row["dias_uteis_entrega_slm"],
                },
            )
        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_crm_fui")
    return render(request, "chatapp/import/import_crm_fui.html")


################################################# IMPORTAR RESPOSTA######################################################


@login_required(login_url='/login/')
def import_resposta(request):
    if request.method == "POST":
        file = request.FILES["file"]
        if not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_respostas_nps")

        df = pd.read_excel(file)
        for _, row in df.iterrows():
            try:
                escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                Respostas_NPS.objects.create(
                    escola=escola,
                    nome=row["nome"],
                    data_da_resposta=row["data_da_resposta"],
                    questao=row["questao"],
                    nota=row["nota"],
                    comentario=row["comentario"]
                )
            except CRM_FUI.DoesNotExist:
                messages.error(
                    request, f"Escola com id_escola {row['id_escola']} não encontrada."
                )
                continue
        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_respostas_nps")
    return render(request, "chatapp/import/import_resposta.html")


################################################# IMPORTAR VENDAS 2024######################################################

@login_required(login_url='/login/')
def import_vendas_slm_2024(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file or not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_vendas_slm_2024")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo Excel: {e}")
            return redirect("import_vendas_slm_2024")

        with transaction.atomic():
            Vendas_SLM_2024.objects.all().delete()  # Limpar a tabela antes de importar novos dados

            for _, row in df.iterrows():
                try:
                    escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                    Vendas_SLM_2024.objects.create(
                        escola=escola,
                        numero_do_pedido=row["numero_do_pedido"],
                        nome_pais=row["nome_pais"],
                        nome_do_aluno=row["nome_do_aluno"],
                        data_do_pedido=row["data_do_pedido"],
                        quantidade=row["quantidade"],
                        id_linha=row["id_linha"],
                    )
                except CRM_FUI.DoesNotExist:
                    messages.error(
                        request, f"Escola com id_escola {row['id_escola']} não encontrada."
                    )
                    continue
                except Exception as e:
                    messages.error(
                        request,
                        f"Erro ao importar a linha com id_escola {row['id_escola']}: {e}",
                    )
                    continue

        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_vendas_slm_2024")
    return render(request, "chatapp/import/import_vendas_slm_2024.html")


################################################# IMPORTAR VENDAS 2025######################################################


@login_required(login_url='/login/')
def import_vendas_slm_2025(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file or not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_vendas_slm_2025")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo Excel: {e}")
            return redirect("import_vendas_slm_2025")

        with transaction.atomic():
            Vendas_SLM_2025.objects.all().delete()  # Limpar a tabela antes de importar novos dados

            for _, row in df.iterrows():
                try:
                    escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                    Vendas_SLM_2025.objects.create(
                        escola=escola,
                        numero_do_pedido=row["numero_do_pedido"],
                        nome_pais=row["nome_pais"],
                        nome_do_aluno=row["nome_do_aluno"],
                        data_do_pedido=row["data_do_pedido"],
                        quantidade=row["quantidade"],
                        id_linha=row["id_linha"],
                    )
                except CRM_FUI.DoesNotExist:
                    messages.error(
                        request, f"Escola com id_escola {row['id_escola']} não encontrada."
                    )
                    continue
                except Exception as e:
                    messages.error(
                        request,
                        f"Erro ao importar a linha com id_escola {row['id_escola']}: {e}",
                    )
                    continue

        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_vendas_slm_2025")
    return render(request, "chatapp/import/import_vendas_slm_2025.html")

################################################# IMPORTAR CLIENTE OCULTO 2024######################################################


@login_required(login_url='/login/')
def import_cliente_oculto_2024(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file or not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_cliente_oculto_2024")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo Excel: {e}")
            return redirect("import_cliente_oculto_2024")

        with transaction.atomic():
            Avaliacao_Cliente_Oculto_24.objects.all().delete()  # Limpar a tabela antes de importar novos dados

            for _, row in df.iterrows():
                try:
                    escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                    Avaliacao_Cliente_Oculto_24.objects.create(
                        escola=escola,
                        categoria=row["categoria"],
                        pergunta=row["pergunta"],
                        resposta=row["resposta"],
                    )
                except CRM_FUI.DoesNotExist:
                    messages.error(
                        request, f"Escola com id_escola {row['id_escola']} não encontrada."
                    )
                    continue
                except Exception as e:
                    messages.error(
                        request,
                        f"Erro ao importar a linha com id_escola {row['id_escola']}: {e}",
                    )
                    continue

        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_cliente_oculto_2024")
    return render(request, "chatapp/import/import_cliente_oculto_2024.html")

################################################# IMPORTAR VISITA ESCOLAS######################################################



@login_required(login_url='/login/')
def import_visitas_escola(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file or not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_visitas_escola")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo Excel: {e}")
            return redirect("import_visitas_escola")

        with transaction.atomic():
            for _, row in df.iterrows():
                try:
                    escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                    
                    # Remove o HTML do campo comentario_visita
                    comentario_limpo = BeautifulSoup(row["comentario_visita"], "html.parser").get_text()

                    Visita_Escola.objects.create(
                        escola=escola,
                        comentario_visita=comentario_limpo
                    )
                except CRM_FUI.DoesNotExist:
                    messages.error(
                        request, f"Escola com id_escola {row['id_escola']} não encontrada."
                    )
                    continue
                except Exception as e:
                    messages.error(
                        request,
                        f"Erro ao importar a linha com id_escola {row['id_escola']}: {e}",
                    )
                    continue

        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_visitas_escola")
    return render(request, "chatapp/import/import_visitas_escola.html")



################################################# IMPORTAR TICKET SPRINKLR######################################################
@login_required(login_url='/login/')
def import_ticket_sprinklr(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file or not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_ticket_sprinklr")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo Excel: {e}")
            return redirect("import_ticket_sprinklr")

        with transaction.atomic():
            Ticket_Sprinklr.objects.all().delete()  # Limpar a tabela antes de importar novos dados

            for _, row in df.iterrows():
                try:
                    escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                    

                    Ticket_Sprinklr.objects.create(
                        escola=escola,
                        id_ticket=row["id_ticket"],
                        cliente=row["cliente"],
                        assunto=row["assunto"],
                        data_ticket=row["data_ticket"],
                    )
                except CRM_FUI.DoesNotExist:
                    messages.error(
                        request, f"Escola com id_escola {row['id_escola']} não encontrada."
                    )
                    continue
                except Exception as e:
                    messages.error(
                        request,
                        f"Erro ao importar a linha com id_escola {row['id_escola']}: {e}",
                    )
                    continue

        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_ticket_sprinklr")
    return render(request, "chatapp/import/import_ticket_sprinklr.html")
############################################# CHAT SAF###########################################################

@login_required(login_url='/login/')
def export_pedidos_alterados_excel(request, school_id):
    # Filtra os pedidos alterados pela escola selecionada
    pedidos = PedidosAlterados.objects.filter(escola__id_escola=school_id)

    # Converte os dados dos pedidos para um DataFrame do pandas
    df = pd.DataFrame(list(pedidos.values('nome_do_aluno', 'numero_do_pedido', 'motivo', 'alterado_por')))

    # Define o nome do arquivo para download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="pedidos_alterados_{school_id}.xlsx"'

    # Exporta o DataFrame para Excel
    df.to_excel(response, index=False)

    return response


def generate_excel_report(vendas):
    df = pd.DataFrame(list(vendas.values()))
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="vendas_slm_2024.xlsx"'
    df.to_excel(response, index=False)


    return response


def escolher_frase_inicial(school):
    frases_iniciais = [
        "Olá, seja bem-vindo ao meu cantinho! Sou o Chinook, o ursinho guia da Maple Bear e vou te mostrar tudo sobre a escola {school.nome_da_escola}.",
        "Bem-vindo! Sou o Chinook, o urso que adora ajudar! Vamos dar uma espiadinha na escola {school.nome_da_escola} da Maple Bear?",
        "E aí! Eu sou o Chinook, o ursinho mais curioso da Maple Bear. Vamos explorar a escola {school.nome_da_escola} juntos?",
        "Oi, eu sou o Chinook, o ursinho que adora compartilhar. Vem comigo conhecer a escola {school.nome_da_escola} da Maple Bear!",
        "Bem-vindo ao meu mundo! Eu sou o Chinook, seu urso guia na Maple Bear e vou te apresentar a escola {school.nome_da_escola} com muito carinho.",
        "Oi, oi! Sou o Chinook, seu amigo de confiança na Maple Bear. Preparado para descobrir tudo sobre a escola {school.nome_da_escola}?",
        "Olá! Eu sou o Chinook, o urso que adora novas aventuras na Maple Bear e hoje vou te levar para conhecer a escola {school.nome_da_escola}.",
        "Oiii, eu sou o Chinook, seu companheiro de pelúcia da Maple Bear, estou animado para te mostrar a escola {school.nome_da_escola}!",
        "Bem-vindo! Eu sou o Chinook, o ursinho mais fofo da Maple Bear, estou aqui para te contar tudo sobre a escola {school.nome_da_escola}.",
        "Olá! Eu sou o Chinook, seu urso amigo na Maple Bear, estou aqui para te guiar pelo universo da escola {school.nome_da_escola}.",
        "Oi! Eu sou o Chinook, seu urso guia na Maple Bear e mal posso esperar para te mostrar tudo sobre a escola {school.nome_da_escola}!",
        "Bem-vindo! Eu sou o Chinook, o urso que adora aventuras e estou aqui para te guiar pela escola {school.nome_da_escola} da Maple Bear.",
        "Oi, oi! Eu sou o Chinook, o ursinho que está sempre por perto na Maple Bear. Vamos juntos conhecer a escola {school.nome_da_escola}?",
        "Oi! Eu sou o Chinook, o ursinho que adora conhecer novos lugares. Vem comigo explorar a escola {school.nome_da_escola}!",
        "Oi, oi! Sou o Chinook, seu amiguinho de aventuras na Maple Bear. Preparado para explorar a escola {school.nome_da_escola}?",
        "Oi! Eu sou o Chinook, seu urso de confiança na Maple Bear. Juntos, vamos conhecer a escola {school.nome_da_escola}!",
        "Oi! Eu sou o Chinook, seu companheiro de todas as horas na Maple Bear. Vamos conhecer juntos a escola {school.nome_da_escola}?",
        "Bem-vindo! Sou o Chinook, o ursinho que adora fazer novos amigos. Hoje, vou te guiar pela escola {school.nome_da_escola} da Maple Bear.",
        "Olá! Eu sou o Chinook, o ursinho sempre pronto para te ajudar. Vamos descobrir juntos tudo sobre a escola {school.nome_da_escola} da Maple Bear!",
        "Oi, oi! Eu sou o Chinook, o urso que adora fazer parte das suas aventuras. Vem comigo conhecer a escola {school.nome_da_escola}!",
        "Bem-vindo! Eu sou o Chinook, o ursinho que sempre tem uma novidade para compartilhar. Hoje, vou te mostrar a escola {school.nome_da_escola}.",
        "Olá! Eu sou o Chinook, seu ursinho guia na Maple Bear. Pronto para explorar a escola {school.nome_da_escola} comigo?",
        "Oi! Eu sou o Chinook, o ursinho que ama ajudar os amigos. Vem comigo descobrir a escola {school.nome_da_escola} da Maple Bear!",
        "Oi, oi! Eu sou o Chinook, o ursinho que adora novas descobertas. Vamos juntos conhecer a escola {school.nome_da_escola}?",
        "Olá! Eu sou o Chinook, o ursinho que adora guiar novas aventuras. Vamos conhecer a escola {school.nome_da_escola} da Maple Bear?",
        "Oi! Eu sou o Chinook, o ursinho sempre animado para novas aventuras. Vamos explorar a escola {school.nome_da_escola} juntos?",
        "Oi, oi! Eu sou o Chinook, seu amiguinho de todas as horas. Preparado para descobrir tudo sobre a escola {school.nome_da_escola} da Maple Bear?",
        "Bem-vindo! Eu sou o Chinook, o ursinho que adora estar por perto. Vem comigo conhecer a escola {school.nome_da_escola}!",
        "Oi, oi! Eu sou o Chinook, o ursinho que ama conhecer novos amigos. Vamos juntos explorar a escola {school.nome_da_escola}?",
        "Bem-vindo! Eu sou o Chinook, o urso que adora guiar aventuras. Hoje, vou te mostrar tudo sobre a escola {school.nome_da_escola}!",
        "Olá! Eu sou o Chinook, o ursinho que está sempre pronto para te ajudar. Vamos conhecer a escola {school.nome_da_escola}?",
        "Oi! Eu sou o Chinook, o urso que adora conhecer novos lugares. Hoje, vou te mostrar tudo sobre a escola {school.nome_da_escola}!",
        "Bem-vindo! Eu sou o Chinook, o ursinho que adora estar ao seu lado. Vamos juntos explorar a escola {school.nome_da_escola}!",

    ]

    return random.choice(frases_iniciais).format(school=school)


def escolher_frase_final(school):
    frases_finais = [
        "Espero ter ajudado! Se precisar de mais alguma coisa, é só me chamar!",
        "Estou sempre por aqui para te ajudar. Qualquer dúvida, conte comigo!",
        "Se precisar de mais informações, estou só a um enter de distância!",
        "Se precisar de mais alguma informação, estarei sempre aqui para te ajudar!",
        "Foi ótimo estar com você! Se precisar de mim, é só chamar.",
        "Se precisar de mais alguma coisa, estou aqui para te ajudar!",
        "Espero que tenha sido útil! Estou por aqui se precisar de mais informações.",
        "Conte comigo sempre que precisar! Estou aqui para ajudar.",
        "Se precisar de mais alguma coisa, estou sempre por aqui!",
        "Estou sempre à disposição! Se precisar, é só me chamar!",
    ]

    return random.choice(frases_finais).format(school=school)




############################################# CHAT SAF###########################################################
#################################################################################################################
#################################################################################################################




@login_required(login_url='/login/')
def filtered_chat_view(request):
    if request.method == "POST":
        data = json.loads(request.body)
        school_id = data.get("school_id")
        message = data.get("message")

        user = request.user

        if not user.api_key:
            return JsonResponse({"error": "Usuário não possui chave API válida"}, status=403)

        api_key = user.api_key

        print("Recebido POST request")
        print(f"ID da escola: {school_id}")
        print(f"Mensagem: {message}")

        if not school_id:
            print("Erro: ID da escola não fornecido")
            return JsonResponse({"error": "School ID not provided"}, status=400)

        # Tenta obter a escola, independentemente do tipo de mensagem
        try:
            school = CRM_FUI.objects.get(id_escola=school_id)
            print(f"Escola encontrada: {school.nome_da_escola}")
        except CRM_FUI.DoesNotExist:
            print("Erro: Escola não encontrada")
            return JsonResponse({"error": "School not found"}, status=404)

        # Busca o resumo do NPS, se existir
        resumo_nps = Resumo_Respostas_NPS.objects.filter(escola=school).first()
        resumo_nps_text = resumo_nps.resumo if resumo_nps else None

        # Converte o texto markdown para HTML se existir
        if resumo_nps_text:
            resumo_nps_text = markdown.markdown(resumo_nps_text, extensions=['nl2br', 'extra'])

            # Substitui <strong> por <span style='font-weight: bold;'>
            resumo_nps_text = resumo_nps_text.replace("<strong>", "<span style='font-weight: bold;'>").replace("</strong>", "</span>")

        # Busca o resumo do NPS, se existir
        resumo_co24 = Resumo_Respostas_ClienteOculto24.objects.filter(escola=school).first()
        resumo_resumo_co24_text = resumo_co24.resumo if resumo_co24 else None

        # Converte o texto markdown para HTML se existir
        if resumo_resumo_co24_text:
            resumo_resumo_co24_text = markdown.markdown(resumo_resumo_co24_text, extensions=['nl2br', 'extra'])

            # Substitui <strong> por <span style='font-weight: bold;'>
            resumo_resumo_co24_text = resumo_resumo_co24_text.replace("<strong>", "<span style='font-weight: bold;'>").replace("</strong>", "</span>")

        # Busca o resumo da Visita da Escola, se existir
        resumo_visita = Resumo_Visita_Escola.objects.filter(escola=school).first()
        resumo_visita_escola = resumo_visita.resumo if resumo_visita else None
        planificador_responses_auto = Planificador_2024.objects.filter(escola__id_escola=school_id).first()  # Pegando o primeiro objeto

        # Converte o texto markdown para HTML se existir
        if resumo_visita_escola: 
            resumo_visita_escola = markdown.markdown(resumo_visita_escola, extensions=['nl2br', 'extra'])

            # Substitui <strong> por <span style='font-weight: bold;'>
            resumo_visita_escola = resumo_visita_escola.replace("<strong>", "<span style='font-weight: bold;'>").replace("</strong>", "</span>")




        if message == 'auto':
            # Registra a requisição com tokens = 0, pois não utiliza a API da OpenAI
            log, created = UserRequestLog.objects.get_or_create(user=user)
            log.request_count += 1
            log.tokens_used += 0  # Não houve uso de tokens
            log.save()


            # Registra uma linha na model RegistroIA para mensagens automáticas
            RegistroIA.objects.create(
                usuario=user,
                escola=school,
                pergunta="Por favor passe o resumo da escola",
                resposta="Resumo gerado",
                tokens_used=0
            )


            complemento = ""
            if school.complemento_escola and school.complemento_escola.lower() not in ["", "null", "nan", "0", "-"]:
                complemento = f" {school.complemento_escola}"
            segmento_info = ""
            if school.segmento_da_escola and school.segmento_da_escola.lower() not in ["", "null", "nan", "em implantação"] \
                    and school.atual_serie and school.atual_serie.lower() not in ["", "null", "nan", "em implantação"]:
                segmento_info = (
                    f"<span style='font-weight: bold;'>Segmento:</span> {school.segmento_da_escola} - "
                    f"<span style='font-weight: bold;'>Atual Série:</span> {school.atual_serie} - "
                    f"<span style='font-weight: bold;'>Avanço de Segmento:</span> {school.avanco_segmento}.<br>"
                )

            inadimplencia_info = ""
            if school.status_de_adimplencia.lower() != "adimplente":
                inadimplencia_info = f" - <span style='font-weight: bold;'>Inadimplência:</span> R$ {school.inadimplencia}"

            # URLs de download
            url_slm_2024 = f"/download_excel_report/?school_id={school_id}"
            url_slm_2025 = f"/download_excel_report_25/?school_id={school_id}"
            url_pedidos_alterados = f"/export_pedidos_alterados/{school_id}/"

            download_icon = '<i class="ri-file-excel-line"  style="font-size: 18px"></i>'
            list_icon = '<i class="ri-feedback-fill"  style="font-size: 18px"></i>'


            frase_inicial = escolher_frase_inicial(school)
            frase_final = escolher_frase_final(school)

            response = (
                f"{frase_inicial}<br><br>"
                
                f"<span style='font-weight: bold;'>Informações Básicas:</span><br>"
                f"<span style='font-weight: bold;'>Id da Escola:</span> {school.id_escola} - <span style='font-weight: bold;'>CNPJ:</span> {school.CNPJ} - <span style='font-weight: bold;'>Cluster:</span> {school.cluster} - <span style='font-weight: bold;'>Status:</span> {school.status_da_escola}.<br>"
                f"<span style='font-weight: bold;'>Endereço:</span> {school.endereco}{complemento}, {school.bairro_escola}, {school.cidade_da_escola}, {school.estado_da_escola}, CEP {school.cep_escola}, na região {school.regiao_da_escola} do Brasil."
                f" - <span style='font-weight: bold;'>Telefone:</span> {school.telefone_de_contato_da_escola} - <span style='font-weight: bold;'>Email:</span> {school.email_da_escola}.<br>"
                f"{segmento_info}"
                
                f"<br><span style='font-weight: bold;'>Vendas e Metas de SLM:</span><br>"
                f"<span style='font-weight: bold;'>SLMs Vendidos 2024:</span> {school.slms_vendidos} - "
                f"<a href='{url_slm_2024}'>{download_icon}</a> - "
                f"<span style='font-weight: bold;'>Meta de SLMs 2024:</span> {school.meta} - "
                f"<span style='font-weight: bold;'>SLMs Vendidos 2025:</span> {school.slms_vendidos_25} - "
                f"<a href='{url_slm_2025}'>{download_icon}</a> - "
                f"<span style='font-weight: bold;'>Meta de SLMs 2025:</span> Ainda não foi definido - "
                f"<span style='font-weight: bold;'>Relatório de Pedidos Cancelados:</span> - "
                f"<a href='{url_pedidos_alterados}'>{download_icon}</a> - "
                f"<span style='font-weight: bold;'>Dias Úteis para Entrega do SLM:</span> {school.dias_uteis_entrega_slm}.<br>"
                
                f"<br><span style='font-weight: bold;'>Avaliações:</span><br>"
                f"<span style='font-weight: bold;'>NPS Pais 2024 - 1ª Onda:</span> {school.nps_pais_2024_1_onda} - "
                f"<span style='font-weight: bold;'>Cliente Oculto 2024:</span> {school.cliente_oculto_2024} - "
                f"<span style='font-weight: bold;'>Quality Assurance 2024: </span> {school.quality_assurance_2024}.<br>"

                f"<br><span style='font-weight: bold;'>Financeiro:</span><br>"
                f"<span style='font-weight: bold;'>Ticket Médio:</span> R$ {school.ticket_medio} - "
                f"<span style='font-weight: bold;'>Valor Royalties:</span> R$ {school.valor_royalties} - "
                f"<span style='font-weight: bold;'>Valor de FDMP:</span> R$ {school.valor_fdmp} - "
                f"<span style='font-weight: bold;'>Status de Adimplência/Inadimplência:</span> {school.status_de_adimplencia}{inadimplencia_info}.<br>"

                f"<br><span style='font-weight: bold;'>Consultores:</span><br>"
                f"<span style='font-weight: bold;'>Consultor Comercial:</span> {school.consultor_comercial} - "
                f"<span style='font-weight: bold;'>Consultor de Gestão Escolar:</span> {school.consultor_gestao_escolar} - "
                f"<span style='font-weight: bold;'>Consultor Acadêmico:</span> {school.consultor_academico} - "
                f"<span style='font-weight: bold;'>Consultor SAF:</span> {school.consultor_saf}.<br>"

                f"<br><span style='font-weight: bold;'>Planificador:</span><br>"
                f"<span style='font-weight: bold;'>Tem CRM B2C Implementado ?</span> {planificador_responses_auto.crm_b2c} - "
                f"<span style='font-weight: bold;'>Tem Árvore Implementado ?</span> {planificador_responses_auto.arvore} - "
                f"<span style='font-weight: bold;'>Tem Toddle Implementado ?</span> {planificador_responses_auto.toddle} - "
                f"<span style='font-weight: bold;'>Circular de Oferta 2025 foi Publicada ?</span> {planificador_responses_auto.circular_oferta_2025_publicado}.<br>"

            )

            # Adiciona o resumo do NPS se estiver disponível
            if resumo_nps_text:
                response += (
                f"<br><span style='font-weight: bold;'>Resumo Comentários Negativos NPS 1° Onda:</span> - "
                f'<a class="open-modal-button" '
                f'data-modal-content="{resumo_nps_text}">'
                f'{list_icon} Ver Resumo</a>'
                )
            # Adiciona o resumo do NPS se estiver disponível
            if resumo_resumo_co24_text:
                response += (
                f"<br><span style='font-weight: bold;'>Resumo Comentários Negativos Cliente Oculo 23/24:</span> - "
                f"<a class='open-modal-button' "
                f"data-modal-content='{resumo_resumo_co24_text}'>"
                f"{list_icon} Ver Resumo</a><br>"
                )
            # Adiciona o resumo do NPS se estiver disponível
            if resumo_visita_escola:
                response += (
                    f"<br><span style='font-weight: bold;'>Resumo da Visita na Escola:</span><br>"
                    f"{resumo_visita_escola}<br>"
                )


            response += f"<br>{frase_final}"


            return JsonResponse({"response": response})


        # Apenas chama a função de classificação se a mensagem não for automática
        question_type = classify_question_chat_central(message, api_key)
        print(f"Tipo de pergunta: {question_type}")

        if question_type == "nps":
            print("Lidando com categoria NPS")
            nps_responses = (
                Respostas_NPS.objects.filter(escola__id_escola=school_id)
                .exclude(comentario__isnull=True)
                .exclude(comentario__exact="")
                .exclude(comentario__exact="nan")
            )
            context = ""
            for response in nps_responses:
                context += (
                    f"NPS: {school.nps_pais_2024_1_onda} - Pontuação Geral dessa Escola.\n"
                    f"Nome do respondente: {response.nome}\n"
                    f"Questão perguntada no NPS: {response.questao}\n"
                    f"Nota: {response.nota} - "
                    f"As notas variam de 1 a 5, exceto para a pergunta de recomendação, que varia de 0 a 10.\n"
                    f"Comentário: {response.comentario}\n\n"
                )
            print("Contexto NPS gerado")




        elif question_type == "ouvidoria":
            print("Lidando com categoria Ouvidoria")

            # Parte 1: Passar resumos das outras escolas primeiro
            outras_escolas_resumos = Resumo_SAC.objects.exclude(escola__id_escola=school_id)
            
            context = "Resumo dos comentários e reclamações de outras escolas:\n"
            for resumo in outras_escolas_resumos:
                context += (
                    f"Escola: {resumo.escola.nome_da_escola}\n"
                    f"Resumo: {resumo.resumo}\n\n"
                )

            # Enviar o contexto de resumos para a API
            print("Contexto inicial (somente resumos) enviado para a API")

            # A API agora vai retornar uma lista das escolas relevantes, baseadas nos resumos.
            escolas_relevantes_ids = processar_resumos_sac(context, api_key)
            
            # Parte 2: Se a API identificar escolas relacionadas ao assunto, busque os detalhes completos de Ouvidoria_SAC
            if escolas_relevantes_ids:
                escolas_relevantes = Ouvidoria_SAC.objects.filter(
                    escola__id_escola__in=escolas_relevantes_ids
                ).exclude(comentario__isnull=True).exclude(comentario__exact="").exclude(comentario__exact="nan")
                
                detalhes_contexto = "Detalhes completos das escolas relacionadas sobre Ouvidoria/SAC:\n"
                for response in escolas_relevantes:
                    detalhes_contexto += (
                        f"Escola: {response.escola.nome_da_escola}\n"
                        f"Comentário: {response.comentario}\n"
                        f"Data: {response.data_reclamacao}\n\n"
                    )

                # Enviar os detalhes das escolas relacionadas para a segunda validação pela API
                escolas_validadas_ids = validar_detalhes_ouvidoria(detalhes_contexto, api_key)

                # Agora incluir no contexto final apenas as escolas que passaram na segunda validação
                if escolas_validadas_ids:
                    context += "\nInformações detalhadas das escolas relacionadas (após validação) da Ouvidoria/SAC:\n"
                    escolas_validadas = Ouvidoria_SAC.objects.filter(
                        escola__id_escola__in=escolas_validadas_ids
                    ).exclude(comentario__isnull=True).exclude(comentario__exact="").exclude(comentario__exact="nan")
                    
                    for response in escolas_validadas:
                        context += (
                            f"Escola: {response.escola.nome_da_escola}\n"
                            f"Comentário: {response.comentario}\n"
                            f"Data: {response.data_reclamacao}\n\n"
                        )

            # Parte 3: Agora adicionamos as informações detalhadas da escola atual
            ouvidoria_responses = (
                Ouvidoria_SAC.objects.filter(escola__id_escola=school_id)
                .exclude(comentario__isnull=True)
                .exclude(comentario__exact="")
                .exclude(comentario__exact="nan")
            )

            context += "\nSituação da escola atual referente à pergunta do usuário sobre Ouvidoria/SAC:\n"
            for response in ouvidoria_responses:
                context += (
                    f"Escola: {school.nome_da_escola}\n"
                    f"Comentário: {response.comentario}\n"
                    f"Data: {response.data_reclamacao}\n\n"
                )

            print("Contexto final gerado com detalhes das escolas relacionadas e da escola atual")

        elif question_type == "pedido":
            print("Categoria identificada: Pedido")

            try:
                print("Extraindo número do pedido da mensagem...")
                order_number = extract_order_number(message, api_key)
                print(f"Número do pedido extraído: {order_number}")
            except Exception as e:
                print(f"Erro ao extrair número do pedido: {e}")
                return JsonResponse({"error": f"Erro ao extrair número do pedido: {e}"}, status=500)

            try:
                print(f"Consultando o pedido {order_number} na base de dados...")
                pedido = Pedido.objects.get(pedido=order_number)
                print(f"Pedido encontrado: {order_number}")
                response = (
                    f"Informações do pedido {order_number}<br>"
                    f"<span style='font-weight: bold;'>Nome da Escola:</span> {pedido.escola.nome_da_escola}<br>"
                    f"<span style='font-weight: bold;'>Data do Pedido:</span> {pedido.data_do_pedido}<br>"
                    f"<span style='font-weight: bold;'>Nome do Produto:</span> {pedido.nome_do_produto}<br>"
                    f"<span style='font-weight: bold;'>Nome do Pai/Mãe:</span> {pedido.nome_pais}<br>"
                    f"<span style='font-weight: bold;'>Nome do Aluno/Aluna:</span> {pedido.nome_do_aluno}<br>"
                    f"<span style='font-weight: bold;'>Código do Produto:</span> {pedido.codigo_do_produto}<br>"
                    f"<span style='font-weight: bold;'>Status ERP:</span> {pedido.status_erp}<br>"
                    f"<span style='font-weight: bold;'>Status Logístico:</span> {pedido.status_logistico}<br>"
                    f"<span style='font-weight: bold;'>Status Transportadora:</span> {pedido.status_transportadora}<br>"
                )
            except Pedido.DoesNotExist:
                print(f"Pedido {order_number} não encontrado.")
                response = f"Pedido {order_number} não encontrado."
            except Exception as e:
                print(f"Erro ao consultar o pedido {order_number}: {e}")
                response = f"Erro ao consultar o pedido: {e}"

            return JsonResponse({"response": response})

        elif question_type == "cliente_oculto":
            print("Lidando com categoria Cliente Oculto")
            co24_responses = (
                Avaliacao_Cliente_Oculto_24.objects.filter(escola__id_escola=school_id)
            )
            context = ""
            context += (
                f"Cliente Oculto é uma avaliação realizada por cliente secreto enviado pela franqueada Maple Bear onde o objetivo é avaliar a escola do ponte de vista de um cliente:\n"
            )
            for response in co24_responses:
                context += (
                    f"Categoria: {response.categoria}\n"
                    f"Questão perguntada no Cliente Oculto 2024: {response.pergunta}\n"
                    f"Resposta: {response.resposta}\n\n"
                )
            print("Contexto Cliente Oculto")

        elif question_type == "planificador":
            print("Lidando com categoria Planificador")
            planificador_responses = Planificador_2024.objects.filter(escola__id_escola=school_id).first()  # Pegando o primeiro objeto

            if planificador_responses:
                context = (
                    f"Informações de controle da escola:\n"
                    f"Nome da Escola: {school.nome_da_escola}\n"
                    f"Bloco Drivers Comerciais Estratégicos:\n"
                    f"Data da última atualização do bloc drivers comerciais estratégicos: {planificador_responses.ultima_data_atualizacao_bloc_drivers_comerciais_estrategicos}\n"
                    f"Tem o sistema CRM B2C nessa escola?: {planificador_responses.crm_b2c}\n"
                    f"Data de Abertura da Matrícula 2025: {planificador_responses.data_abertura_matricula_2025}\n"
                    f"Teve a Circular de Oferta 2025 Publicada nessa escola?: {planificador_responses.circular_oferta_2025_publicado}\n"
                    f"Data de Abertura da Circular 2025 caso tenha: {planificador_responses.data_de_abertura_da_circular_2025}\n"
                    f"Tem o sistema Toddle nessa escola?: {planificador_responses.toddle}\n"
                    f"Tem o sistema Árvore nessa escola?: {planificador_responses.arvore}\n"
                    f"Data de Implementação da Árvore caso tenha: {planificador_responses.data_implementacao_arvore}\n"

                    f"Bloco Funil Comercial:\n"
                    f"Última Data de Atualização do Bloco Funil Comercial: {planificador_responses.ultima_data_atualizacao_bloc_funil_comercial}\n"
                    f"Leads Central Ago 24: {planificador_responses.leads_central_ago_24}\n"
                    f"Leads Escolas Ago 24: {planificador_responses.leads_escolas_ago_24}\n"
                    f"Visitas Ago 24: {planificador_responses.visitas_ago_24}\n"
                    f"Taxa de Conversão Atual Leads Visitas: {planificador_responses.taxa_conversao_atual_leads_visitas}\n"
                    f"Matrículas Ago 24: {planificador_responses.matriculas_ago_24}\n"
                    f"Taxa de Conversão Atual Visitas Matrículas: {planificador_responses.taxa_conversao_atual_visitas_matriculas}\n"
                    f"Taxa de Conversão Leads Matrículas: {planificador_responses.taxa_conversao_leads_matriculas}\n"

                    f"Bloco Drivers Comerciais Meio:\n"
                    f"Última Data de Atualização do Bloc Drivers Comerciais Meio: {planificador_responses.ultima_data_atualizacao_bloc_drivers_comerciais_meio}\n"
                    f"Meta Alunos 5K 2024: {planificador_responses.meta_alunos_5k_2024}\n"
                    f"Setup Plano Comercial Segundo Semestre: {planificador_responses.setup_plano_comercial_segundo_semestre}\n"
                    f"Essa escola já fez a Ação 1 Elegível Trade Marketing?: {planificador_responses.acao_1_elegivel_trade_marketing}\n"
                    f"Essa escola já fez a Ação 1 Trade Valor?: {planificador_responses.acao_1_trade_valor}\n"
                    f"Essa escola já fez a Ação 1 Trade Marketing Ações Alinhadas?: {planificador_responses.acao_1_trade_marketing_acoes_alinhadas}\n"
                    f"Essa escola já fez a Ação 2 Experience Day 10/08/24?: {planificador_responses.acao_2_experience_day_10_08_24}\n"
                    f"Essa escola já fez a Ação 2 Experience Day 24/08/24?: {planificador_responses.acao_2_experience_day_24_08_24}\n"
                    f"Essa escola já fez a Ação 2 Experience Day 21/09/24?: {planificador_responses.acao_2_experience_day_21_09_24}\n"
                    f"Essa escola já fez a Ação 2 Experience Day 26/10/24?: {planificador_responses.acao_2_experience_day_26_10_24}\n"
                    f"Essa escola já fez a Ação 2 Experience Day 09/11/24?: {planificador_responses.acao_2_experience_day_09_11_24}\n"
                    f"Essa escola já fez a Ação 3 Friend Get Friend?: {planificador_responses.acao_3_friend_get_friend}\n"
                    f"Essa escola já fez a Ação 4 Webinars com Autoridades Pré?: {planificador_responses.acao_4_webinars_com_autoridades_pre}\n"
                    f"Essa escola já fez a Ação 4 Webinars com Autoridades Pós?: {planificador_responses.acao_4_webinars_com_autoridades_pos}\n"
                    f"Essa escola já fez a Piloto Welcome Baby Bear?: {planificador_responses.piloto_welcome_baby_bear}\n"
                    f"Ação 5 SDR Taxa de Conversão Validação Lead: {planificador_responses.acao_5_sdr_taxa_conversao_validacao_lead}\n"
                    f"Ação 5 SDR Taxa de Conversão Visitas: {planificador_responses.acao_5_sdr_taxa_conversao_visitas}\n"
                    f"Ação 6 Alinhado Resgate Leads: {planificador_responses.acao_6_alinhado_resgate_leads}\n"
                    f"Ação 6 Quantidade de Leads Resgatados: {planificador_responses.acao_6_quantidade_leads_resgatados}\n"
                    f"Ação 6 Todos Leads Resgatados Contatados: {planificador_responses.acao_6_todos_leads_resgatados_contatados}\n"
                    f"Data de Atualização dos Resultados: {planificador_responses.data_atualizacao_resultados}\n"
                    f"SLM 2022: {planificador_responses.slm_2022}\n"
                    f"SLM 2023: {planificador_responses.slm_2023}\n"
                    f"Meta Orçamentária 2024: {planificador_responses.meta_orcamentaria_2024}\n"
                    f"Base Rematriculáveis 2025: {planificador_responses.base_rematriculaveis_2025}\n"
                    f"Meta Rematrícula 2025: {planificador_responses.meta_rematricula_2025}\n"
                    f"Real Rematrículas 2025: {planificador_responses.real_rematriculas_2025}\n"
                    f"Atingimento Rematrículas 2025: {planificador_responses.atingimento_rematriculas_2025}\n"
                    f"Meta Matrícula 2025: {planificador_responses.meta_matricula_2025}\n"
                    f"Real Matrícula 2025: {planificador_responses.real_matricula_2025}\n"
                    f"Atingimento Matrículas 2025: {planificador_responses.atingimento_matriculas_2025}\n"
                    f"Total Meta Alunos 2025: {planificador_responses.total_meta_alunos_2025}\n"
                    f"Total Real Alunos 2025: {planificador_responses.total_real_alunos_2025}\n"
                    f"Atingimento Real Alunos 2025: {planificador_responses.atingimento_real_alunos_2025}\n"
                    f"Correlação Alunos SLMS 2025: {planificador_responses.correlacao_alunos_slms_2025}\n"
                    f"MC 2025: {planificador_responses.mc_2025}\n"
                    f"SLMS 2025 M: {planificador_responses.slms_2025_m}\n"
                    f"Pedidos Represados Logística 2025: {planificador_responses.pedidos_represados_logistica_2025}\n"
                    f"Pedidos Faturados: {planificador_responses.pedidos_faturados}\n"
                    f"Pedidos Entregues: {planificador_responses.pedidos_entregues}\n"
                )
            else:
                context = "Não há informações disponíveis sobre o Planificador para esta escola.\n"

            print("Contexto Planificador gerado")

        elif question_type == "sprinklr":
            print("Lidando com categoria Sprinklr")
            sprinklr_responses = Ticket_Sprinklr.objects.filter(escola__id_escola=school_id)

            # Contagem total de tickets distintos
            total_tickets = sprinklr_responses.values('id_ticket').distinct().count()

            # Ranking de assuntos (mantido sem ajuste, pois é baseado em contagem total)
            assuntos_ranking = (
                sprinklr_responses
                .values('assunto')
                .annotate(total=Count('assunto'))
                .order_by('-total')
            )

            # Ranking de clientes com contagem distinta de tickets
            clientes_ranking = (
                sprinklr_responses
                .values('cliente')
                .annotate(total=Count('id_ticket', distinct=True))
                .order_by('-total')
            )

            # Contagem de tickets por data_ticket com contagem distinta
            data_ranking = (
                sprinklr_responses
                .values('data_ticket')
                .annotate(total=Count('id_ticket', distinct=True))
                .order_by('data_ticket')
            )

            context = (
                f"Sprinklr é o sistema de atendimentos por tickets da Maple Bear:\n\n"
                f"Resumo das informações gerais da Sprinklr dessa escola:\n"
                f"Total de Tickets Distintos: {total_tickets}\n\n"
            )

            context += "\nRanking de Assuntos:\n"

            # Adiciona o ranking de assuntos ao contexto
            for assunto in assuntos_ranking:
                context += f"{assunto['assunto']} - Total de Tickets: {assunto['total']}\n"

            context += "\nTotal de Tickets por Data:\n"

            # Adiciona a contagem de tickets por data_ticket ao contexto
            for data in data_ranking:
                context += f"{data['data_ticket']} - Total de Tickets: {data['total']}\n"

            context += "\nRanking de Clientes:\n"

            # Adiciona o ranking de clientes ao contexto
            for cliente in clientes_ranking:
                context += f"{cliente['cliente']} - Total de Tickets: {cliente['total']}\n"

            context += "\nDetalhe dos Tickets:\n"

            for response in sprinklr_responses:
                context += (
                    f"["
                    f"Id do Ticket: {response.id_ticket}\n"
                    f"Cliente: {response.cliente}\n"
                    f"Assunto: {response.assunto}\n"
                    f"Data: {response.data_ticket}\n"
                    f"]\n"
                )

            context += "\nImportante detalhe responda apenas com as informações solicitadas pelo o usuario, estou te passando todo esse contexto para facilitar o fornecimento das informações. Outra coisa caso ele peça um resumo, tente resumir tudo em uma paragrafo apenas explicando em texto, ou seja, veja o contexto e interprete a informação e forneça uma analise do que foi pedido não copie e contexto apenas cole na resposta.\n"

            print("Contexto Sprinklr gerado")


        elif question_type == "analise completa da escola":
            print("Lidando com análise completa da escola")
            
            # Obter respostas do NPS relacionadas à escola
            nps_responses = (
                Respostas_NPS.objects.filter(escola__id_escola=school_id)
                .exclude(comentario__isnull=True)
                .exclude(comentario__exact="")
                .exclude(comentario__exact="nan")
            )
            
            # Construir o contexto inicial com informações básicas da escola
            context = (
                f"Informações Básicas da Escola:\n"
                f"Nome da Escola: {school.nome_da_escola}\n"
                f"Status: {school.status_da_escola}\n"
                f"Cluster: {school.cluster}\n"
                f"CEP: {school.cep_escola}\n"
                f"Endereço: {school.endereco}\n"
                f"Complemento: {school.complemento_escola}\n"
                f"Bairro: {school.bairro_escola}\n"
                f"Cidade: {school.cidade_da_escola}\n"
                f"Estado: {school.estado_da_escola}\n"
                f"Região: {school.regiao_da_escola}\n"
                f"Segmento: {school.segmento_da_escola}\n"
                f"Atual Série: {school.atual_serie}\n"
                f"Avanço Segmento: {school.avanco_segmento}\n"
                f"Vendas e Metas de SLM:\n"
                f"SLMs Vendidos 2024: {school.slms_vendidos} - Referente ao ano de 2024.\n"
                f"SLMs Vendidos 2025: {school.slms_vendidos_25} - Referente ao ano de 2025.\n"
                f"Meta de SLMs 2024: {school.meta} - Meta de vendas para 2024.\n"
                f"Meta de SLMs 2025: Ainda não definida para 2025.\n"
                f"Avaliações:\n"
                f"NPS Pais 2024 - 1° Onda: {school.nps_pais_2024_1_onda} - Pontuação do NPS (Net Promoter Score) dos pais dos alunos no primeiro semestre de 2024.\n"
                f"Cliente Oculto 2024: {school.cliente_oculto_2024} - Pontuação da avaliação feita por cliente oculto em 2024.\n"
                f"Quality Assurance 2024: {school.quality_assurance_2024} - Pontuação da avaliação de qualidade da escola em 2024.\n"
                f"Financeiro:\n"
                f"Ticket Médio: {school.ticket_medio} - Valor médio de mensalidade.\n"
                f"Valor Royalties: {school.valor_royalties} - Valor mensal de royalties devido à franqueada.\n"
                f"Valor de FDMP (Fundo de Marketing): {school.valor_fdmp} - Valor mensal de contribuição ao fundo de marketing.\n"
                f"Status de Adimplência/Inadimplência: {school.status_de_adimplencia}.\n"
            )
            
            if school.status_de_adimplencia == "Inadimplente":
                context += f"Inadimplência: R$ {school.inadimplencia} - Valor devido pela escola.\n"
            
            # Adicionar respostas do NPS ao contexto
            if nps_responses.exists():
                context += "\nRespostas do NPS dessa escola:\n"
                for response in nps_responses:
                    context += (
                        f"Questão perguntada no NPS: {response.questao}\n"
                        f"Comentário: {response.comentario}\n\n"
                    )
            else:
                context += "\nNenhuma resposta do NPS encontrada para esta escola.\n"
            
            # Parte 1: Obter os resumos de todas as entradas da base de conhecimento
            resumos = Base_de_Conhecimento_Geral.objects.values("id", "resumo")
            if not resumos:
                print("Nenhum resumo encontrado na base de conhecimento.")
                context += "\nBase de Conhecimento:\nNenhum resumo encontrado na base de conhecimento.\n"
            else:
                # Formatar o contexto para enviar para a API
                context_resumos = "\n".join([f"ID: {r['id']} - Resumo: {r['resumo']}" for r in resumos if r['resumo']])
                if not context_resumos.strip():
                    print("Nenhum resumo válido para enviar à API.")
                    context += "\nBase de Conhecimento:\nNenhum resumo válido para enviar à base de conhecimento.\n"
                else:
                    print("Enviando resumos para a API para identificar IDs relevantes...")
                    
                    # Parte 2: Filtrar os IDs relevantes com base na pergunta do usuário
                    try:
                        ids_relevantes = filtrar_resumos_conhecimento(message, context_resumos, api_key)
                        print(f"IDs relevantes recebidos: {ids_relevantes}")
                    except Exception as e:
                        print(f"Erro ao processar IDs relevantes: {str(e)}")
                        return JsonResponse({"error": "Erro ao processar IDs relevantes."}, status=500)

                    # Parte 3: Construir o contexto com os resumos relevantes
                    if not ids_relevantes:
                        print("Nenhum ID relevante foi retornado pela API.")
                        context += "\nBase de Conhecimento:\nNenhum ID relevante encontrado na base de conhecimento.\n"
                    else:
                        entradas_relevantes = Base_de_Conhecimento_Geral.objects.filter(id__in=ids_relevantes)
                        context += "\nBase de Conhecimento:\n"
                        for entrada in entradas_relevantes:
                            context += (
                                f"Título: {entrada.titulo}\n"
                                f"Assunto: {entrada.topico}\n"
                                f"Sub Assunto: {entrada.sub_topico}\n"
                                f"Resumo: {entrada.resumo}\n\n"
                            )

            print("Contexto de análise completa da escola gerado")
        
        elif question_type == "base de conhecimento":
            print("Lidando com categoria Base de Conhecimento")
            
            # Parte 1: Obter os resumos de todas as entradas
            resumos = Base_de_Conhecimento_Geral.objects.values("id", "resumo")
            if not resumos:
                print("Nenhum resumo encontrado na base de conhecimento.")
                return JsonResponse({"error": "Nenhum resumo encontrado."}, status=404)

            # Formatar o contexto para enviar para a API
            context_resumos = "\n".join([f"ID: {r['id']} - Resumo: {r['resumo']}" for r in resumos if r['resumo']])
            if not context_resumos.strip():
                print("Nenhum resumo válido para enviar à API.")
                return JsonResponse({"error": "Nenhum resumo válido para enviar."}, status=404)

            print("Enviando resumos para a API para identificar IDs relevantes...")
            
            # Parte 2: Filtrar os IDs relevantes com base na pergunta do usuário
            try:
                ids_relevantes = filtrar_resumos_conhecimento(message, context_resumos, api_key)
                print(f"IDs relevantes recebidos: {ids_relevantes}")
            except Exception as e:
                print(f"Erro ao processar IDs relevantes: {str(e)}")
                return JsonResponse({"error": "Erro ao processar IDs relevantes."}, status=500)

            # Parte 3: Construir o contexto com os resumos relevantes
            if not ids_relevantes:
                print("Nenhum ID relevante foi retornado pela API.")
                return JsonResponse({"error": "Nenhum ID relevante encontrado."}, status=404)

            entradas_relevantes = Base_de_Conhecimento_Geral.objects.filter(id__in=ids_relevantes)
            context = "Entradas relevantes encontradas:\n"
            for entrada in entradas_relevantes:
                context += (
                    f"Título: {entrada.titulo}\n"
                    f"Assunto: {entrada.topico}\n"
                    f"Sub Assunto: {entrada.sub_topico}\n"
                    f"Resumo: {entrada.resumo}\n\n"
                )

            print("Contexto gerado com base de conhecimento relevante")
        
        elif question_type in ["vendas", "relatório de vendas"]:
            vendas_responses_2024 = Vendas_SLM_2024.objects.filter(
                escola__id_escola=school_id
            )
            total_quantidade_2024 = vendas_responses_2024.aggregate(total=Sum('quantidade'))['total']

            vendas_responses_2025 = Vendas_SLM_2025.objects.filter(
                escola__id_escola=school_id
            )
            total_quantidade_2025 = vendas_responses_2025.aggregate(total=Sum('quantidade'))['total']

            context = (
                f"O total de vendas da escola no ciclo de 2024 foi {total_quantidade_2024} e no ciclo de 2025 foi {total_quantidade_2025}. "
                f"Para outras informações, você pode ver o relatório completo em Excel das vendas do ciclo de 2024 clicando "
                f"[aqui](/download_excel_report/?school_id={school_id}) e para as vendas do ciclo de 2025 clicando "
                f"[aqui](/download_excel_report_25/?school_id={school_id})."
            )
            print("Contexto de relatório de vendas gerado")
            
        else:
            context = (
                f"Informações Básicas da Escola:\n"
                f"Nome da Escola: {school.nome_da_escola}\n"
                f"CNPJ: {school.CNPJ}\n"
                f"Status: {school.status_da_escola}\n"
                f"Cluster: {school.cluster}\n"
                f"CEP: {school.cep_escola}\n"
                f"Endereço: {school.endereco}\n"
                f"Complemento: {school.complemento_escola}\n"
                f"Bairro: {school.bairro_escola}\n"
                f"Cidade: {school.cidade_da_escola}\n"
                f"Estado: {school.estado_da_escola}\n"
                f"Região: {school.regiao_da_escola}\n"
                f"Telefone: {school.telefone_de_contato_da_escola}\n"
                f"Email: {school.email_da_escola}\n"
                f"Segmento: {school.segmento_da_escola}\n"
                f"Atual Série: {school.atual_serie}\n"
                f"Avanço Segmento: {school.avanco_segmento}\n"
                
                f"Vendas e Metas de SLM:\n"
                f"SLMs Vendidos 2024: {school.slms_vendidos} - SLM ou SLMs no plural são os materiais vendidos, esses aqui são referente a 2024.\n"
                f"SLMs Vendidos 2025: {school.slms_vendidos_25} - SLM ou SLMs no plural são os materiais vendidos, esses aqui são referente a 2025.\n"
                f"Meta de SLMs 2024: {school.meta} - Esse campo é a meta de Vendas de SLM vendidos.\n"
                f"Meta de SLMs 2025: Ainda não foi definido a meta por escola.\n"
                f"Dias Úteis para Entrega do SLM nessa escola: {school.dias_uteis_entrega_slm} - Basicamente o prazo de entrega do material nessa escola.\n"
                
                f"Avaliações:\n"
                f"NPS Pais 2024 - 1° Onda: {school.nps_pais_2024_1_onda} - "
                f"Este campo indica a pontuação referente ao NPS(Net Promoter Score) dos pais dos alunos que estudam na escola, que foi realizado no 1° semestre no ano(1° Onda).\n"
                f"Cliente Oculto 2024: {school.cliente_oculto_2024} - "
                f"Este campo indica a pontuação referente ao Cliente Oculto, que uma avaliação realizada por uma empresa terceirizada onde consiste em um falso cliente ir até a escola para avaliar ela.\n"
                f"Quality Assurance 2024: {school.quality_assurance_2024} - "
                f"Este campo indica a pontuação referente Quality Assurance uma avaliação realizada para ver a qualidade da escola.\n"
                
                f"Financeiro:\n"
                f"Ticket Médio: {school.ticket_medio} - Este é o valor médio de mensalidade cobrada pela escola.\n"
                f"Valor Royalties: {school.valor_royalties} - Este é o valor de royalties que a escola deve pagar por mês à franqueada Maple Bear.\n"
                f"Valor de FDMP(Fundo de Marketing): {school.valor_fdmp} - Este é o valor de FDMP(Fundo de Marketing) que a escola deve pagar por mês à franqueada Maple Bear.\n"
                f"Status de Adimplência/Inadimplência: {school.status_de_adimplencia} - "
                f"Este campo indica se a escola está Adimplente ou Inadimplente referente aos seus pagamentos(Royalties e FDMP) que devem ser feitos à franqueada Maple Bear.\n"
            )

            if school.status_de_adimplencia == "Inadimplente":
                context += f"Inadimplência: {school.inadimplencia} - Este é o valor que a escola está devendo para a Maple Bear.\n"

            context += (
                f"Consultores:\n"
                f"Consultor Comercial: {school.consultor_comercial}\n"
                f"Consultor Gestão Escolar: {school.consultor_gestao_escolar}\n"
                f"Consultor Acadêmico: {school.consultor_academico}\n"
                f"Consultor SAF(Serviço de Atendimento ao Franqueado): {school.consultor_saf}\n"
            )

        response_data = config_chat_central(message, api_key, context)
        response_html = response_data['formatted_response']
        tokens_used = response_data['tokens']

        response_text = BeautifulSoup(response_html, "html.parser").get_text()

        RegistroIA.objects.create(
            usuario=user,
            escola=school,
            pergunta=message,
            resposta=response_text,
            tokens_used=tokens_used
        )

        log, created = UserRequestLog.objects.get_or_create(user=user)
        log.request_count += 1
        log.tokens_used += tokens_used
        log.save()

        return JsonResponse({"response": response_html})
    else:
        schools = CRM_FUI.objects.all().order_by("nome_da_escola")
        return render(request, "chatapp/filtered_chat.html", {"schools": schools})





############################################# CHAT EXCOM###########################################################
#################################################################################################################
#################################################################################################################

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')


@login_required(login_url='/login/')
def excom(request):
    if request.method == "POST":
        data = json.loads(request.body)
        school_id = data.get("school_id")
        message = data.get("message")

        user = request.user

        if not user.api_key:
            return JsonResponse({"error": "Usuário não possui chave API válida"}, status=403)

        api_key = user.api_key

        print("Recebido POST request")
        print(f"ID da escola: {school_id}")
        print(f"Mensagem: {message}")

        if not school_id:
            print("Erro: ID da escola não fornecido")
            return JsonResponse({"error": "School ID not provided"}, status=400)

        # Tenta obter a escola, independentemente do tipo de mensagem
        try:
            school = CRM_FUI.objects.get(id_escola=school_id)
            print(f"Escola encontrada: {school.nome_da_escola}")
        except CRM_FUI.DoesNotExist:
            print("Erro: Escola não encontrada")
            return JsonResponse({"error": "School not found"}, status=404)

        # Busca o resumo do NPS, se existir
        resumo_nps = Resumo_Respostas_NPS.objects.filter(escola=school).first()
        resumo_nps_text = resumo_nps.resumo if resumo_nps else None

        # Converte o texto markdown para HTML se existir
        if resumo_nps_text:
            resumo_nps_text = markdown.markdown(resumo_nps_text, extensions=['nl2br', 'extra'])

            # Substitui <strong> por <span style='font-weight: bold;'>
            resumo_nps_text = resumo_nps_text.replace("<strong>", "<span style='font-weight: bold;'>").replace("</strong>", "</span>")

        # Busca o resumo do NPS, se existir
        resumo_co24 = Resumo_Respostas_ClienteOculto24.objects.filter(escola=school).first()
        resumo_resumo_co24_text = resumo_co24.resumo if resumo_co24 else None

        # Converte o texto markdown para HTML se existir
        if resumo_resumo_co24_text:
            resumo_resumo_co24_text = markdown.markdown(resumo_resumo_co24_text, extensions=['nl2br', 'extra'])

            # Substitui <strong> por <span style='font-weight: bold;'>
            resumo_resumo_co24_text = resumo_resumo_co24_text.replace("<strong>", "<span style='font-weight: bold;'>").replace("</strong>", "</span>")




        if message == 'auto':

            # Soma dos valores de processos com status específico
            status_ativos = ["ATIVO", "EM GRAU DE RECURSO", "REDISTRIBUIDO", "AUTUADO", "SEGREDO DE JUSTIÇA", "JULGADO", "TRANSITADO EM JULGADO"]
            soma_valores_processos = Processo.objects.filter(escola=school, status_do_processo__in=status_ativos).aggregate(total_valor=Sum('valor'))['total_valor'] or 0

            # Soma dos valores e quantidade dos protestos com resultado 'Positivo'
            soma_valores_protestos = Protesto.objects.filter(escola=school, resultado="Positivo").aggregate(total_valor=Sum('valor'), total_quantidade=Sum('quantidade'))
            total_valor_protestos = soma_valores_protestos['total_valor'] or 0
            total_quantidade_protestos = soma_valores_protestos['total_quantidade'] or 0

            # Formatação dos valores no estilo brasileiro
            soma_valores_processos_formatado = locale.format_string('%.2f', soma_valores_processos, grouping=True)
            total_valor_protestos_formatado = locale.format_string('%.2f', total_valor_protestos, grouping=True)


            # Registra a requisição com tokens = 0, pois não utiliza a API da OpenAI
            log, created = UserRequestLog.objects.get_or_create(user=user)
            log.request_count += 1
            log.tokens_used += 0  # Não houve uso de tokens
            log.save()


            complemento = ""
            if school.complemento_escola and school.complemento_escola.lower() not in ["", "null", "nan", "0", "-"]:
                complemento = f" {school.complemento_escola}"
            segmento_info = ""
            if school.segmento_da_escola and school.segmento_da_escola.lower() not in ["", "null", "nan", "em implantação"] \
                    and school.atual_serie and school.atual_serie.lower() not in ["", "null", "nan", "em implantação"]:
                segmento_info = (
                    f"<span style='font-weight: bold;'>Segmento:</span> {school.segmento_da_escola} - "
                    f"<span style='font-weight: bold;'>Atual Série:</span> {school.atual_serie} - "
                    f"<span style='font-weight: bold;'>Avanço de Segmento:</span> {school.avanco_segmento}.<br>"
                )

            inadimplencia_info = ""
            if school.status_de_adimplencia.lower() != "adimplente":
                inadimplencia_info = f" - <span style='font-weight: bold;'>Inadimplência:</span> R$ {school.inadimplencia}"

            # URLs de download
            url_slm_2024 = f"/download_excel_report/?school_id={school_id}"
            url_slm_2025 = f"/download_excel_report_25/?school_id={school_id}"
            url_pedidos_alterados = f"/export_pedidos_alterados/{school_id}/"

            download_icon = '<i class="ri-file-excel-line"  style="font-size: 18px"></i>'

            frase_inicial = escolher_frase_inicial(school)
            frase_final = escolher_frase_final(school)

            response = (
                f"{frase_inicial}<br><br>"
                
                f"<span style='font-weight: bold;'>Informações Básicas:</span><br>"
                f"<span style='font-weight: bold;'>Id da Escola:</span> {school.id_escola} - <span style='font-weight: bold;'>CNPJ:</span> {school.CNPJ} - <span style='font-weight: bold;'>Cluster:</span> {school.cluster} - <span style='font-weight: bold;'>Status:</span> {school.status_da_escola}.<br>"
                f"<span style='font-weight: bold;'>Endereço:</span> {school.endereco}{complemento}, {school.bairro_escola}, {school.cidade_da_escola}, {school.estado_da_escola}, CEP {school.cep_escola}, na região {school.regiao_da_escola} do Brasil."
                f" - <span style='font-weight: bold;'>Telefone:</span> {school.telefone_de_contato_da_escola} - <span style='font-weight: bold;'>Email:</span> {school.email_da_escola}.<br>"
                f"{segmento_info}"
                
                f"<br><span style='font-weight: bold;'>Vendas e Metas de SLM:</span><br>"
                f"<span style='font-weight: bold;'>SLMs Vendidos 2024:</span> {school.slms_vendidos} - "
                f"<a href='{url_slm_2024}'>{download_icon}</a> - "
                f"<span style='font-weight: bold;'>Meta de SLMs 2024:</span> {school.meta} - "
                f"<span style='font-weight: bold;'>SLMs Vendidos 2025:</span> {school.slms_vendidos_25} - "
                f"<a href='{url_slm_2025}'>{download_icon}</a> - "
                f"<span style='font-weight: bold;'>Relatório de Pedidos Cancelados:</span> - "
                f"<a href='{url_pedidos_alterados}'>{download_icon}</a> - "
                f"<span style='font-weight: bold;'>Meta de SLMs 2025:</span> Ainda não foi definido - "
                f"<span style='font-weight: bold;'>Dias Úteis para Entrega do SLM:</span> {school.dias_uteis_entrega_slm}.<br>"
                
                f"<br><span style='font-weight: bold;'>Avaliações:</span><br>"
                f"<span style='font-weight: bold;'>NPS Pais 2024 - 1ª Onda:</span> {school.nps_pais_2024_1_onda} - "
                f"<span style='font-weight: bold;'>Cliente Oculto 2024:</span> {school.cliente_oculto_2024} - "
                f"<span style='font-weight: bold;'>Quality Assurance 2024: </span> {school.quality_assurance_2024}.<br>"

                f"<br><span style='font-weight: bold;'>Financeiro:</span><br>"
                f"<span style='font-weight: bold;'>Ticket Médio:</span> R$ {school.ticket_medio} - "
                f"<span style='font-weight: bold;'>Valor Royalties:</span> R$ {school.valor_royalties} - "
                f"<span style='font-weight: bold;'>Valor de FDMP:</span> R$ {school.valor_fdmp} - "
                f"<span style='font-weight: bold;'>Status de Adimplência/Inadimplência:</span> {school.status_de_adimplencia}{inadimplencia_info}.<br>"

                f"<br><span style='font-weight: bold;'>Informações Jurídicas:</span><br>"
                f"<span style='font-weight: bold;'>Valor Total de Processos Ativos:</span> R$ {soma_valores_processos_formatado}<br>"
                f"<span style='font-weight: bold;'>Valor Total de Protestos:</span> Valor Total - R$ {total_valor_protestos_formatado} - Quantidade Total de Protestos - {total_quantidade_protestos}<br>"
            )

            # Adiciona o resumo do NPS se estiver disponível
            if resumo_nps_text:
                response += (
                    f"<br><span style='font-weight: bold;'>Resumo das Respostas Negativas do NPS:</span><br>"
                    f"{resumo_nps_text}<br>"
                )
            # Adiciona o resumo do NPS se estiver disponível
            if resumo_resumo_co24_text:
                response += (
                    f"<br><span style='font-weight: bold;'>Resumo dos Comentários Negativos do Cliente Oculto:</span><br>"
                    f"{resumo_resumo_co24_text}<br>"
                )
            # Adiciona a frase final
            response += f"<br>{frase_final}"

            return JsonResponse({"response": response})


        # Apenas chama a função de classificação se a mensagem não for automática
        question_type = classify_question_excom(message, api_key)
        print(f"Tipo de pergunta: {question_type}")

        if question_type == "nps":
            print("Lidando com categoria NPS")
            nps_responses = (
                Respostas_NPS.objects.filter(escola__id_escola=school_id)
                .exclude(comentario__isnull=True)
                .exclude(comentario__exact="")
                .exclude(comentario__exact="nan")
            )
            context = ""
            for response in nps_responses:
                context += (
                    f"NPS: {school.nps_pais_2024_1_onda} - Pontuação Geral dessa Escola.\n"
                    f"Nome do respondente: {response.nome}\n"
                    f"Questão perguntada no NPS: {response.questao}\n"
                    f"Nota: {response.nota} - "
                    f"As notas variam de 1 a 5, exceto para a pergunta de recomendação, que varia de 0 a 10.\n"
                    f"Comentário: {response.comentario}\n\n"
                )
            print("Contexto NPS gerado")




        elif question_type == "ouvidoria":
            print("Lidando com categoria Ouvidoria")

            # Parte 1: Passar resumos das outras escolas primeiro
            outras_escolas_resumos = Resumo_SAC.objects.exclude(escola__id_escola=school_id)
            
            context = "Resumo dos comentários e reclamações de outras escolas:\n"
            for resumo in outras_escolas_resumos:
                context += (
                    f"Escola: {resumo.escola.nome_da_escola}\n"
                    f"Resumo: {resumo.resumo}\n\n"
                )

            # Enviar o contexto de resumos para a API
            print("Contexto inicial (somente resumos) enviado para a API")

            # A API agora vai retornar uma lista das escolas relevantes, baseadas nos resumos.
            escolas_relevantes_ids = processar_resumos_sac(context, api_key)
            
            # Parte 2: Se a API identificar escolas relacionadas ao assunto, busque os detalhes completos de Ouvidoria_SAC
            if escolas_relevantes_ids:
                escolas_relevantes = Ouvidoria_SAC.objects.filter(
                    escola__id_escola__in=escolas_relevantes_ids
                ).exclude(comentario__isnull=True).exclude(comentario__exact="").exclude(comentario__exact="nan")
                
                detalhes_contexto = "Detalhes completos das escolas relacionadas sobre Ouvidoria/SAC:\n"
                for response in escolas_relevantes:
                    detalhes_contexto += (
                        f"Escola: {response.escola.nome_da_escola}\n"
                        f"Comentário: {response.comentario}\n"
                        f"Data: {response.data_reclamacao}\n\n"
                    )

                # Enviar os detalhes das escolas relacionadas para a segunda validação pela API
                escolas_validadas_ids = validar_detalhes_ouvidoria(detalhes_contexto, api_key)

                # Agora incluir no contexto final apenas as escolas que passaram na segunda validação
                if escolas_validadas_ids:
                    context += "\nInformações detalhadas das escolas relacionadas (após validação) da Ouvidoria/SAC:\n"
                    escolas_validadas = Ouvidoria_SAC.objects.filter(
                        escola__id_escola__in=escolas_validadas_ids
                    ).exclude(comentario__isnull=True).exclude(comentario__exact="").exclude(comentario__exact="nan")
                    
                    for response in escolas_validadas:
                        context += (
                            f"Escola: {response.escola.nome_da_escola}\n"
                            f"Comentário: {response.comentario}\n"
                            f"Data: {response.data_reclamacao}\n\n"
                        )

            # Parte 3: Agora adicionamos as informações detalhadas da escola atual
            ouvidoria_responses = (
                Ouvidoria_SAC.objects.filter(escola__id_escola=school_id)
                .exclude(comentario__isnull=True)
                .exclude(comentario__exact="")
                .exclude(comentario__exact="nan")
            )

            context += "\nSituação da escola atual referente à pergunta do usuário sobre Ouvidoria/SAC:\n"
            for response in ouvidoria_responses:
                context += (
                    f"Escola: {school.nome_da_escola}\n"
                    f"Comentário: {response.comentario}\n"
                    f"Data: {response.data_reclamacao}\n\n"
                )

            print("Contexto final gerado com detalhes das escolas relacionadas e da escola atual")



        elif question_type == "cliente_oculto":
            print("Lidando com categoria Cliente Oculto")
            co24_responses = (
                Avaliacao_Cliente_Oculto_24.objects.filter(escola__id_escola=school_id)
            )
            context = ""
            context += (
                f"Cliente Oculto é uma avaliação realizada por cliente secreto enviado pela franqueada Maple Bear onde o objetivo é avaliar a escola do ponte de vista de um cliente:\n"
            )
            for response in co24_responses:
                context += (
                    f"Categoria: {response.categoria}\n"
                    f"Questão perguntada no Cliente Oculto 2024: {response.pergunta}\n"
                    f"Resposta: {response.resposta}\n\n"
                )
            print("Contexto Cliente Oculto")

        
        else:
            context = (
                f"Nome da Escola: {school.nome_da_escola}\n"
            )

            
        response_data = config_chat_excom(message, api_key, context)

        tokens_used = response_data['tokens']  # Ajuste conforme o retorno da função utils

        # Atualiza o log de requisições com os tokens usados
        log, created = UserRequestLog.objects.get_or_create(user=user)
        log.request_count += 1
        log.tokens_used += tokens_used
        log.save()

        return JsonResponse({"response": response_data['formatted_response']})
    else:
        schools = CRM_FUI.objects.all().order_by(
            "nome_da_escola"
        )  # Ordenar alfabeticamente
        return render(request, "chatapp/excom.html", {"schools": schools})





@login_required(login_url='/login/')
def import_processo_protestos(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        planilha_tipo = request.POST.get("planilha_tipo")

        if not file or not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_data")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo Excel: {e}")
            return redirect("import_data")

        df = df.where(pd.notnull(df), None)

        if planilha_tipo == "processos":
            with transaction.atomic():
                Processo.objects.all().delete()

                for _, row in df.iterrows():
                    try:
                        escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                        data_publicacao = row["data_publicacao"] if pd.notna(row["data_publicacao"]) else None

                        Processo.objects.create(
                            escola=escola,
                            numero_do_processo=row["numero_do_processo"],
                            assunto_principal=row["assunto_principal"],
                            data_publicacao=data_publicacao,
                            valor=row["valor"] if row["valor"] is not None else None,
                            tipo_do_processo=row["tipo_do_processo"],
                            tipo_de_tribunal=row["tipo_de_tribunal"],
                            status_do_processo=row["status_do_processo"],
                            nome_polo_ativo=row["nome_polo_ativo"],
                            documento_polo_ativo=row["documento_polo_ativo"],
                            nome_polo_passivo=row["nome_polo_passivo"],
                            documento_polo_passivo=row["documento_polo_passivo"]
                        )
                    except CRM_FUI.DoesNotExist:
                        messages.error(request, f"Escola com id_escola {row['id_escola']} não encontrada.")
                        continue

        elif planilha_tipo == "protestos":
            with transaction.atomic():
                Protesto.objects.all().delete()

                for _, row in df.iterrows():
                    try:
                        escola = CRM_FUI.objects.get(id_escola=row["id_escola"])

                        quantidade = int(row["quantidade"]) if pd.notna(row["quantidade"]) else None
                        valor = Decimal(row["valor"]) if pd.notna(row["valor"]) else None

                        Protesto.objects.create(
                            escola=escola,
                            documento=row["documento"],
                            status=row["status"],
                            resultado=row["resultado"],
                            cartorio=row["cartorio"],
                            quantidade=quantidade,
                            valor=valor
                        )
                    except CRM_FUI.DoesNotExist:
                        messages.error(request, f"Escola com id_escola {row['id_escola']} não encontrada.")
                        continue
            messages.success(request, "Dados importados com sucesso!")
            return redirect("import_processo_protestos")

    return render(request, "chatapp/import/import_processo_protestos.html")


#################################################################################################################
#################################################################################################################
#################################################################################################################






def download_excel_report_slm_2024(request):
    school_id = request.GET.get("school_id")
    vendas_responses = Vendas_SLM_2024.objects.filter(escola__id_escola=school_id)
    df = pd.DataFrame(list(vendas_responses.values()))
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="vendas_slm_2024_{school_id}.xlsx"'
    )
    df.to_excel(response, index=False)
    return response



def download_excel_report_slm_2025(request):
    school_id = request.GET.get("school_id")
    vendas_responses = Vendas_SLM_2025.objects.filter(escola__id_escola=school_id)
    df = pd.DataFrame(list(vendas_responses.values()))
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="vendas_slm_2025_{school_id}.xlsx"'
    )
    df.to_excel(response, index=False)
    return response


################################### Controle Escolas #####################################################


class ControleEscolasSearchView(ListView):
    model = CRM_FUI
    template_name = "chatapp/controle_escolas.html"
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "")
        context["order_by"] = self.request.GET.get("order_by", "nome_da_escola")
        page_obj = context["page_obj"]

        # Obtém o número da página atual
        current_page = page_obj.number

        # Se há mais de 5 páginas
        if page_obj.paginator.num_pages > 5:
            if current_page - 2 < 1:
                start_page = 1
                end_page = 5
            elif current_page + 2 > page_obj.paginator.num_pages:
                start_page = page_obj.paginator.num_pages - 4
                end_page = page_obj.paginator.num_pages
            else:
                start_page = current_page - 2
                end_page = current_page + 2
        else:
            start_page = 1
            end_page = page_obj.paginator.num_pages

        context["page_range"] = range(start_page, end_page + 1)

        # Adiciona porcentagem_planificador ao contexto
        for escola in context['object_list']:
            planificador = Planificador_2024.objects.filter(escola=escola).first()
            if planificador:
                escola.planificador = planificador
                escola.porcentagem_planificador = calcular_porcentagem_sim(planificador)
            else:
                escola.planificador = None
                escola.porcentagem_planificador = 0

        return context

    def get_queryset(self):
        query = self.request.GET.get("q")
        order_by = self.request.GET.get("order_by", "nome_da_escola")
        if query:
            return CRM_FUI.objects.filter(Q(nome_da_escola__icontains=query)).order_by(order_by)
        return CRM_FUI.objects.all().order_by(order_by)
    


################################### Gerar Resumo NPS#####################################################

def gerar_resumo_nps(request, school_id):
    print("Iniciando a geração do resumo NPS...")

    # Busca a escola pelo ID
    escola = get_object_or_404(CRM_FUI, id_escola=school_id)
    print(f"Escola encontrada: {escola.nome_da_escola}")

    # Filtra as respostas NPS para a escola, excluindo comentários vazios, null ou nan
    respostas = Respostas_NPS.objects.filter(
        escola=escola
    ).exclude(
        comentario__isnull=True
    ).exclude(
        comentario__exact=""
    ).exclude(
        comentario__iexact="nan"
    )
    print(f"Número de respostas encontradas: {respostas.count()}")

    # Cria um contexto categorizado com as questões e comentários para passar para o prompt
    comentarios_categorizados = "\n".join(
        [f"Categoria: {resposta.questao}\nComentário: {resposta.comentario}\n" for resposta in respostas]
    )

    if comentarios_categorizados:
        prompt = (
            "Faça um resumo bem resumido dos comentários negativos, monte um paragrafo unico resumindo:\n"
            f"{comentarios_categorizados}"
        )
        print(f"Prompt gerado: {prompt[:100]}...")  # Exibe apenas os primeiros 100 caracteres do prompt

        api_key = request.user.api_key  # Assume que o usuário logado tem uma chave API
        resumo = config_resumo_nps(prompt, api_key)
        print(f"Resumo gerado: {resumo[:100]}...")  # Exibe apenas os primeiros 100 caracteres do resumo

        # Salva o resumo na model Resumo_Respostas_NPS
        resumo_nps, created = Resumo_Respostas_NPS.objects.get_or_create(escola=escola)
        resumo_nps.resumo = resumo
        resumo_nps.save()
        print("Resumo salvo com sucesso!")

        return redirect('controle_escolas')  # Redireciona de volta para a lista de escolas

    print("Não há comentários válidos para resumir.")
    return redirect('controle_escolas')


def gerar_resumos_todas_escolas(request):
    escolas = CRM_FUI.objects.all()
    resultados = []
    
    for escola in escolas:
        # Verifica se já existe um resumo para a escola e se está em branco
        resumo_nps, created = Resumo_Respostas_NPS.objects.get_or_create(escola=escola)
        if resumo_nps.resumo is None or resumo_nps.resumo.strip() == "":
            # Filtra as respostas NPS para a escola, excluindo comentários vazios, null ou nan
            respostas = Respostas_NPS.objects.filter(
                escola=escola
            ).exclude(
                comentario__isnull=True
            ).exclude(
                comentario__exact=""
            ).exclude(
                comentario__iexact="nan"
            )

            comentarios_categorizados = "\n".join(
                [f"Categoria: {resposta.questao}\nComentário: {resposta.comentario}\n" for resposta in respostas]
            )

            if comentarios_categorizados:
                prompt = (
                    "Faça um resumo bem resumido dos comentários negativos, monte um paragrafo unico resumindo:\n"
                    f"{comentarios_categorizados}"
                )
                api_key = request.user.api_key  # Assume que o usuário logado tem uma chave API
                resumo = config_resumo_nps(prompt, api_key)

                # Salva o resumo na model Resumo_Respostas_NPS
                resumo_nps.resumo = resumo
                resumo_nps.save()

                resultados.append(f"Resumo gerado para {escola.nome_da_escola} com sucesso!")
            else:
                resultados.append(f"Não há comentários válidos para {escola.nome_da_escola}.")
        else:
            resultados.append(f"O resumo para {escola.nome_da_escola} já existe e não está em branco.")

        time.sleep(5)

    return JsonResponse({"resultados": resultados})



### Resumo comentarios positivos e negativos ###
def gerar_resumos_todas_escolas_geral(request):
    escolas = CRM_FUI.objects.all()
    resultados = []
    
    for escola in escolas:
        # Verifica se já existe um resumo para a escola e se está em branco
        resumo_nps, created = Resumo_Respostas_NPS_1_Onda_Geral.objects.get_or_create(escola=escola)
        if resumo_nps.resumo is None or resumo_nps.resumo.strip() == "":
            # Filtra as respostas NPS para a escola, excluindo comentários vazios, null ou nan
            respostas = Respostas_NPS.objects.filter(
                escola=escola
            ).exclude(
                comentario__isnull=True
            ).exclude(
                comentario__exact=""
            ).exclude(
                comentario__iexact="nan"
            )

            comentarios_categorizados = "\n".join(
                [f"Categoria: {resposta.questao}\nComentário: {resposta.comentario}\n" for resposta in respostas]
            )

            if comentarios_categorizados:
                prompt = (
                    "Faça um resumo bem resumido dos comentários, monte um paragrafo unico resumindo:\n"
                    f"{comentarios_categorizados}"
                )
                api_key = request.user.api_key  # Assume que o usuário logado tem uma chave API
                resumo = config_resumo_nps(prompt, api_key)

                # Salva o resumo na model Resumo_Respostas_NPS
                resumo_nps.resumo = resumo
                resumo_nps.save()

                resultados.append(f"Resumo gerado para {escola.nome_da_escola} com sucesso!")
            else:
                resultados.append(f"Não há comentários válidos para {escola.nome_da_escola}.")
        else:
            resultados.append(f"O resumo para {escola.nome_da_escola} já existe e não está em branco.")

        time.sleep(5)

    return JsonResponse({"resultados": resultados})





################################### SAC/Ouvidoria #####################################################

@login_required(login_url='/login/')
def import_ouvidoria_sac(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file or not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_ouvidoria")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo Excel: {e}")
            return redirect("import_ouvidoria")

        with transaction.atomic():
            Ouvidoria_SAC.objects.all().delete()  # Limpar a tabela antes de importar novos dados

            for _, row in df.iterrows():
                try:
                    escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                    

                    Ouvidoria_SAC.objects.create(
                        escola=escola,
                        origem=row["origem"],
                        tema=row["tema"],
                        nome_responsavel=row["nome_responsavel"],
                        comentario=row["comentario"],
                        data_reclamacao=row["data_reclamacao"],
                    )
                except CRM_FUI.DoesNotExist:
                    messages.error(
                        request, f"Escola com id_escola {row['id_escola']} não encontrada."
                    )
                    continue
                except Exception as e:
                    messages.error(
                        request,
                        f"Erro ao importar a linha com id_escola {row['id_escola']}: {e}",
                    )
                    continue

        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_ouvidoria")
    return render(request, "chatapp/import/import_ouvidoria.html")




def gerar_resumos_todas_escolas_sac(request):
    escolas = CRM_FUI.objects.all()
    resultados = []
    
    for escola in escolas:
        # Verifica se já existe um resumo para a escola e se está em branco
        resumo_sac, created = Resumo_SAC.objects.get_or_create(escola=escola)
        if resumo_sac.resumo is None or resumo_sac.resumo.strip() == "":
            # Filtra as respostas NPS para a escola, excluindo comentários vazios, null ou nan
            respostas = Ouvidoria_SAC.objects.filter(
                escola=escola
            ).exclude(
                comentario__isnull=True
            ).exclude(
                comentario__exact=""
            ).exclude(
                comentario__iexact="nan"
            )

            comentarios_categorizados = "\n".join(
                [f"Comentário: {resposta.comentario}\n" for resposta in respostas]
            )

            if comentarios_categorizados:
                prompt = (
                    "Faça um resumo bem resumido dos comentários, monte um paragrafo unico resumindo todos comentários desses atendimentos realizados pelo o SAC:\n"
                    f"{comentarios_categorizados}"
                )
                api_key = request.user.api_key  # Assume que o usuário logado tem uma chave API
                resumo = config_resumo_sac(prompt, api_key)

                # Salva o resumo na model Resumo_Respostas_NPS
                resumo_sac.resumo = resumo
                resumo_sac.save()

                resultados.append(f"Resumo gerado para {escola.nome_da_escola} com sucesso!")
            else:
                resultados.append(f"Não há comentários válidos para {escola.nome_da_escola}.")
        else:
            resultados.append(f"O resumo para {escola.nome_da_escola} já existe e não está em branco.")

        time.sleep(5)

    return JsonResponse({"resultados": resultados})



################################### Gerar Resumo Cliente Oculto 2024#####################################################


def gerar_resumo_cliente_oculto(request, school_id):
    print("Iniciando a geração do resumo do Cliente Oculto 2024...")

    # Busca a escola pelo ID
    escola = get_object_or_404(CRM_FUI, id_escola=school_id)
    print(f"Escola encontrada: {escola.nome_da_escola}")

    # Filtra as respostas do Cliente Oculto para a escola, excluindo respostas vazias, null ou nan
    respostas = Avaliacao_Cliente_Oculto_24.objects.filter(
        escola=escola
    ).exclude(
        resposta__isnull=True
    ).exclude(
        resposta__exact=""
    ).exclude(
        resposta__iexact="nan"
    )
    print(f"Número de respostas encontradas: {respostas.count()}")

    # Cria um contexto categorizado com as questões e respostas para passar para o prompt
    respostas_categorizadas = "\n".join(
        [f"Categoria: {resposta.categoria}\nPergunta: {resposta.pergunta}\nResposta: {resposta.resposta}\n" for resposta in respostas]
    )

    if respostas_categorizadas:
        prompt = (
            "Faça um resumo bem resumido das respostas, monte um paragrafo único resumindo:\n"
            f"{respostas_categorizadas}"
        )
        print(f"Prompt gerado: {prompt[:100]}...")  # Exibe apenas os primeiros 100 caracteres do prompt

        api_key = request.user.api_key  # Assume que o usuário logado tem uma chave API
        resumo = config_resumo_cliente_oculto(prompt, api_key)
        print(f"Resumo gerado: {resumo[:100]}...")  # Exibe apenas os primeiros 100 caracteres do resumo

        # Salva o resumo na model Resumo_Respostas_ClienteOculto24
        resumo_co24, created = Resumo_Respostas_ClienteOculto24.objects.get_or_create(escola=escola)
        resumo_co24.resumo = resumo
        resumo_co24.save()
        print("Resumo salvo com sucesso!")

        return redirect('controle_escolas')  # Redireciona de volta para a lista de escolas

    print("Não há respostas válidas para resumir.")
    return redirect('controle_escolas')


def gerar_resumos_cliente_oculto_todas_escolas(request):
    escolas = CRM_FUI.objects.all()
    resultados = []
    
    for escola in escolas:
        # Verifica se já existe um resumo para a escola e se está em branco
        resumo_co24, created = Resumo_Respostas_ClienteOculto24.objects.get_or_create(escola=escola)
        if resumo_co24.resumo is None or resumo_co24.resumo.strip() == "":
            # Filtra as respostas do Cliente Oculto para a escola, excluindo respostas vazias, null ou nan
            respostas = Avaliacao_Cliente_Oculto_24.objects.filter(
                escola=escola
            ).exclude(
                resposta__isnull=True
            ).exclude(
                resposta__exact=""
            ).exclude(
                resposta__iexact="nan"
            )

            respostas_categorizadas = "\n".join(
                [f"Categoria: {resposta.categoria}\nPergunta: {resposta.pergunta}\nResposta: {resposta.resposta}\n" for resposta in respostas]
            )

            if respostas_categorizadas:
                prompt = (
                    "Faça um resumo bem resumido das respostas, monte um paragrafo único resumindo:\n"
                    f"{respostas_categorizadas}"
                )
                api_key = request.user.api_key  # Assume que o usuário logado tem uma chave API
                resumo = config_resumo_cliente_oculto(prompt, api_key)

                # Salva o resumo na model Resumo_Respostas_ClienteOculto24
                resumo_co24.resumo = resumo
                resumo_co24.save()

                resultados.append(f"Resumo gerado para {escola.nome_da_escola} com sucesso!")
            else:
                resultados.append(f"Não há respostas válidas para {escola.nome_da_escola}.")
        else:
            resultados.append(f"O resumo para {escola.nome_da_escola} já existe e não está em branco.")

        time.sleep(5)  # Pausa de 5 segundos entre as escolas

    return JsonResponse({"resultados": resultados})




###################################GERAR RESUMO VISITA ESCOLAS#####################################################


def gerar_resumo_visita_escola(request, visita_id):
    print("Iniciando a geração do resumo da Visita Escola...")

    # Busca a visita pelo ID
    visita = get_object_or_404(Visita_Escola, id=visita_id)
    print(f"Visita encontrada: {visita}")

    # Busca a escola associada à visita
    escola = visita.escola
    print(f"Escola associada: {escola.nome_da_escola}")

    # Cria o prompt para gerar o resumo da visita
    prompt = (
        "Faça um resumo bem resumido do comentário da visita, monte um parágrafo único resumindo:\n"
        f"Comentário: {visita.comentario_visita}"
    )
    print(f"Prompt gerado: {prompt[:100]}...")  # Exibe apenas os primeiros 100 caracteres do prompt

    api_key = request.user.api_key  # Assume que o usuário logado tem uma chave API
    resumo = config_resumo_visita_escola(prompt, api_key)
    print(f"Resumo gerado: {resumo[:100]}...")  # Exibe apenas os primeiros 100 caracteres do resumo

    # Salva o resumo na model Resumo_Visita_Escola
    resumo_visita, created = Resumo_Visita_Escola.objects.get_or_create(
        escola=escola,
        visita=visita
    )
    resumo_visita.resumo = resumo
    resumo_visita.save()
    print("Resumo salvo com sucesso!")

    return redirect('controle_visitas') 


def gerar_resumos_visita_escola_todas(request):
    visitas = Visita_Escola.objects.all()
    resultados = []

    for visita in visitas:
        escola = visita.escola
        resumo_visita, created = Resumo_Visita_Escola.objects.get_or_create(
            escola=escola,
            visita=visita
        )

        # Verifica se já existe um resumo e se está em branco
        if resumo_visita.resumo is None or resumo_visita.resumo.strip() == "":
            prompt = (
                "Faça um resumo bem resumido do comentário da visita, monte um parágrafo único resumindo:\n"
                f"Comentário: {visita.comentario_visita}"
            )
            api_key = request.user.api_key  # Assume que o usuário logado tem uma chave API
            resumo = config_resumo_visita_escola(prompt, api_key)

            # Salva o resumo na model Resumo_Visita_Escola
            resumo_visita.resumo = resumo
            resumo_visita.save()

            resultados.append(f"Resumo gerado para visita da escola {escola.nome_da_escola} com sucesso!")
        else:
            resultados.append(f"O resumo para visita da escola {escola.nome_da_escola} já existe e não está em branco.")

        time.sleep(5)  # Pausa de 5 segundos entre as visitas

    return JsonResponse({"resultados": resultados})

################################### PLANIFICADOR #####################################################

class ImportPlanificadorView(LoginRequiredMixin, View):
    login_url = '/login/'

    def get(self, request):
        return render(request, "chatapp/import/import_planificador.html")

    def post(self, request):
        file = request.FILES["file"]
        if not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_planificador")

        df = pd.read_excel(file)
        for _, row in df.iterrows():
            try:
                escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                Planificador_2024.objects.update_or_create(
                    escola=escola,
                    defaults={
                        "ultima_data_atualizacao_bloc_drivers_comerciais_estrategicos": self.parse_date(row["ultima_data_atualizacao_bloc_drivers_comerciais_estrategicos"]),
                        "crm_b2c": self.parse_text(row["crm_b2c"]),
                        "data_abertura_matricula_2025": self.parse_date(row["data_abertura_matricula_2025"]),
                        "circular_oferta_2025_publicado": self.parse_text(row["circular_oferta_2025_publicado"]),
                        "data_de_abertura_da_circular_2025": self.parse_date(row["data_de_abertura_da_circular_2025"]),
                        "toddle": self.parse_text(row["toddle"]),
                        "toddle_planejamento": self.parse_text(row["toddle_planejamento"]),
                        "toddle_portfolio": self.parse_text(row["toddle_portfolio"]),
                        "arvore": self.parse_text(row["arvore"]),
                        "data_implementacao_arvore": self.parse_date(row["data_implementacao_arvore"]),
                        "ultima_data_atualizacao_bloc_funil_comercial": self.parse_date(row["ultima_data_atualizacao_bloc_funil_comercial"]),
                        "leads_central_ago_24": self.parse_number(row["leads_central_ago_24"]),
                        "leads_escolas_ago_24": self.parse_number(row["leads_escolas_ago_24"]),
                        "visitas_ago_24": self.parse_number(row["visitas_ago_24"]),
                        "taxa_conversao_atual_leads_visitas": self.parse_number(row["taxa_conversao_atual_leads_visitas"]),
                        "matriculas_ago_24": self.parse_number(row["matriculas_ago_24"]),
                        "taxa_conversao_atual_visitas_matriculas": self.parse_number(row["taxa_conversao_atual_visitas_matriculas"]),
                        "taxa_conversao_leads_matriculas": self.parse_number(row["taxa_conversao_leads_matriculas"]),
                        "ultima_data_atualizacao_bloc_drivers_comerciais_meio": self.parse_date(row["ultima_data_atualizacao_bloc_drivers_comerciais_meio"]),
                        "meta_alunos_5k_2024": self.parse_number(row["meta_alunos_5k_2024"]),
                        "setup_plano_comercial_segundo_semestre": self.parse_text(row["setup_plano_comercial_segundo_semestre"]),
                        "acao_1_elegivel_trade_marketing": self.parse_text(row["acao_1_elegivel_trade_marketing"]),
                        "acao_1_trade_valor": self.parse_number(row["acao_1_trade_valor"]),
                        "acao_1_trade_marketing_acoes_alinhadas": self.parse_text(row["acao_1_trade_marketing_acoes_alinhadas"]),
                        "acao_2_experience_day_10_08_24": self.parse_text(row["acao_2_experience_day_10_08_24"]),
                        "acao_2_experience_day_24_08_24": self.parse_text(row["acao_2_experience_day_24_08_24"]),
                        "acao_2_experience_day_21_09_24": self.parse_text(row["acao_2_experience_day_21_09_24"]),
                        "acao_2_experience_day_26_10_24": self.parse_text(row["acao_2_experience_day_26_10_24"]),
                        "acao_2_experience_day_09_11_24": self.parse_text(row["acao_2_experience_day_09_11_24"]),
                        "acao_3_friend_get_friend": self.parse_text(row["acao_3_friend_get_friend"]),
                        "acao_4_webinars_com_autoridades_pre": self.parse_text(row["acao_4_webinars_com_autoridades_pre"]),
                        "acao_4_webinars_com_autoridades_pos": self.parse_text(row["acao_4_webinars_com_autoridades_pos"]),
                        "piloto_welcome_baby_bear": self.parse_text(row["piloto_welcome_baby_bear"]),
                        "acao_5_sdr_taxa_conversao_validacao_lead": self.parse_number(row["acao_5_sdr_taxa_conversao_validacao_lead"]),
                        "acao_5_sdr_taxa_conversao_visitas": self.parse_number(row["acao_5_sdr_taxa_conversao_visitas"]),
                        "acao_6_alinhado_resgate_leads": self.parse_text(row["acao_6_alinhado_resgate_leads"]),
                        "acao_6_quantidade_leads_resgatados": self.parse_number(row["acao_6_quantidade_leads_resgatados"]),
                        "acao_6_todos_leads_resgatados_contatados": self.parse_text(row["acao_6_todos_leads_resgatados_contatados"]),
                        "data_atualizacao_resultados": self.parse_date(row["data_atualizacao_resultados"]),
                        "slm_2022": self.parse_number(row["slm_2022"]),
                        "slm_2023": self.parse_number(row["slm_2023"]),
                        "meta_orcamentaria_2024": self.parse_number(row["meta_orcamentaria_2024"]),
                        "base_rematriculaveis_2025": self.parse_number(row["base_rematriculaveis_2025"]),
                        "meta_rematricula_2025": self.parse_number(row["meta_rematricula_2025"]),
                        "real_rematriculas_2025": self.parse_number(row["real_rematriculas_2025"]),
                        "atingimento_rematriculas_2025": self.parse_number(row["atingimento_rematriculas_2025"]),
                        "meta_matricula_2025": self.parse_number(row["meta_matricula_2025"]),
                        "real_matricula_2025": self.parse_number(row["real_matricula_2025"]),
                        "atingimento_matriculas_2025": self.parse_number(row["atingimento_matriculas_2025"]),
                        "total_meta_alunos_2025": self.parse_number(row["total_meta_alunos_2025"]),
                        "total_real_alunos_2025": self.parse_number(row["total_real_alunos_2025"]),
                        "atingimento_real_alunos_2025": self.parse_number(row["atingimento_real_alunos_2025"]),
                        "correlacao_alunos_slms_2025": self.parse_number(row["correlacao_alunos_slms_2025"]),
                        "mc_2025": self.parse_number(row["mc_2025"]),
                        "slms_2025_m": self.parse_number(row["slms_2025_m"]),
                        "pedidos_represados_logistica_2025": self.parse_number(row["pedidos_represados_logistica_2025"]),
                        "pedidos_faturados": self.parse_number(row["pedidos_faturados"]),
                        "pedidos_entregues": self.parse_number(row["pedidos_entregues"]),
                    }
                )
            except CRM_FUI.DoesNotExist:
                messages.error(request, f"Escola com id_escola {row['id_escola']} não encontrada.")
        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_planificador")

    def parse_date(self, date):
        if pd.isna(date):
            return None
        if isinstance(date, str):
            return datetime.fromisoformat(date).date()
        if isinstance(date, pd.Timestamp):
            return date.to_pydatetime().date()
        if isinstance(date, datetime):
            return date.date()
        return date

    def parse_number(self, value):
        if pd.isna(value):
            return None
        return value

    def parse_text(self, value):
        if pd.isna(value):
            return ''
        return value

class PlanificadorCreateView(LoginRequiredMixin, View):
    login_url = '/login/'

    def get(self, request):
        form = PlanificadorForm()
        return render(request, "chatapp/planificador/planificador_form.html", {"form": form})

    def post(self, request):
        form = PlanificadorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Dados criados com sucesso!")
            return redirect("planificador_list")
        return render(request, "chatapp/planificador/planificador_form.html", {"form": form})


def calcular_porcentagem_sim(planificador):
    sim_nao_fields = [
        planificador.crm_b2c,
        planificador.circular_oferta_2025_publicado,
        planificador.toddle,
        planificador.arvore,
        planificador.setup_plano_comercial_segundo_semestre,
        planificador.acao_2_experience_day_10_08_24,
        planificador.acao_2_experience_day_24_08_24,
        planificador.acao_2_experience_day_21_09_24,
        planificador.acao_2_experience_day_26_10_24,
        planificador.acao_2_experience_day_09_11_24,
        planificador.acao_4_webinars_com_autoridades_pre,
        planificador.acao_4_webinars_com_autoridades_pos,
        planificador.piloto_welcome_baby_bear,
        planificador.acao_6_alinhado_resgate_leads,
        planificador.acao_6_todos_leads_resgatados_contatados
    ]

    total_fields = len(sim_nao_fields)
    sim_count = sum(1 for field in sim_nao_fields if field == 'SIM')
    porcentagem_sim = (sim_count / total_fields) * 100
    return f"{porcentagem_sim:.2f}"

def registrar_alteracoes(planificador_antigo, planificador_novo, form, usuario):
    alteracoes = []
    
    # Verifica as mudanças diretamente no formulário
    for campo in form.changed_data:
        valor_antigo = form.initial.get(campo)
        valor_novo = form.cleaned_data.get(campo)
        alteracoes.append(f"{campo}: '{valor_antigo}' -> '{valor_novo}'")

    # Print para verificar se encontrou alguma alteração
    print(f"Alterações encontradas: {alteracoes}")

    # Se houver alterações, salva no histórico
    if alteracoes:
        HistoricoAlteracoes.objects.create(
            planificador=planificador_novo,
            usuario=usuario,
            data_alteracao=timezone.now(),
            alteracoes="\n".join(alteracoes)
        )
        print(f"Alterações salvas no histórico para o usuário {usuario}")
    else:
        print("Nenhuma alteração foi encontrada.")



class PlanificadorUpdateView(LoginRequiredMixin, View):
    login_url = '/login/'

    def get(self, request, pk):
        planificador = get_object_or_404(Planificador_2024, pk=pk)
        form = PlanificadorForm(instance=planificador)

        # Definindo os campos como readonly
        form.fields['slm_2022'].widget.attrs['readonly'] = True
        form.fields['slm_2023'].widget.attrs['readonly'] = True

        porcentagem_planificador = calcular_porcentagem_sim(planificador)

        escola = planificador.escola
        context = {
            "form": form,
            "planificador": planificador,
            "crm_fui_meta": escola.meta,
            "crm_fui_slms_vendidos": escola.slms_vendidos,
            "crm_fui_slms_vendidos_25": escola.slms_vendidos_25,
            "crm_fui_id_escola": escola.id_escola,
            "crm_fui_nome_da_escola": escola.nome_da_escola,
            "crm_fui_cluster": escola.cluster,
            "crm_fui_atual_serie": escola.atual_serie,
            "crm_fui_segmento_da_escola": escola.segmento_da_escola,
            "crm_fui_consultor_comercial": escola.consultor_comercial,
            "crm_fui_consultor_gestao_escolar": escola.consultor_gestao_escolar,
            "crm_fui_consultor_saf": escola.consultor_saf,
            "crm_fui_consultor_academico": escola.consultor_academico,
            "porcentagem_planificador": porcentagem_planificador
        }
        return render(request, "chatapp/planificador/planificador_form_edit.html", context)

    def post(self, request, pk):
        planificador = get_object_or_404(Planificador_2024, pk=pk)
        data = request.POST.copy()
        data['escola'] = planificador.escola.pk  # Inclua o campo escola nos dados do formulário
        form = PlanificadorForm(data, instance=planificador)

        # Definindo os campos como readonly na postagem também
        form.fields['slm_2022'].widget.attrs['readonly'] = True
        form.fields['slm_2023'].widget.attrs['readonly'] = True

        if form.is_valid():
            try:
                # Guarda o estado antigo antes de salvar
                planificador_antigo = Planificador_2024.objects.get(pk=planificador.pk)
                
                planificador_atualizado = form.save(commit=False)
                planificador_atualizado.usuario_modificacao = request.user
                planificador_atualizado.save()

                # Print para verificar se o planificador foi salvo corretamente
                print(f"Planificador {planificador_atualizado.id} salvo com sucesso.")

                # Função para registrar as alterações no histórico
                registrar_alteracoes(planificador_antigo, planificador_atualizado, form, request.user)

                messages.success(request, "Dados atualizados com sucesso!")
                return redirect("buscar_escolas")
            except Exception as e:
                messages.error(request, "Erro ao atualizar os dados.")
                print(f"Erro ao salvar o planificador: {e}")
        else:
            messages.error(request, "Erro ao atualizar os dados: %s" % form.errors)
            print(f"Erros de validação: {form.errors}")

        porcentagem_planificador = calcular_porcentagem_sim(planificador)

        escola = planificador.escola
        context = {
            "form": form,
            "planificador": planificador,
            "crm_fui_meta": escola.meta,
            "crm_fui_slms_vendidos": escola.slms_vendidos,
            "crm_fui_slms_vendidos_25": escola.slms_vendidos_25,
            "crm_fui_id_escola": escola.id_escola,
            "crm_fui_nome_da_escola": escola.nome_da_escola,
            "crm_fui_cluster": escola.cluster,
            "crm_fui_atual_serie": escola.atual_serie,
            "crm_fui_segmento_da_escola": escola.segmento_da_escola,
            "crm_fui_consultor_comercial": escola.consultor_comercial,
            "crm_fui_consultor_gestao_escolar": escola.consultor_gestao_escolar,
            "crm_fui_consultor_saf": escola.consultor_saf,
            "crm_fui_consultor_academico": escola.consultor_academico,
            "porcentagem_planificador": porcentagem_planificador
        }
        return render(request, "chatapp/planificador/planificador_form_edit.html", context)


@csrf_exempt
def gerar_resumo_alteracoes_json(request):
    if request.method == 'POST':
        try:
            # Tenta carregar o corpo do request, mas se estiver vazio, ignora.
            if request.body:
                dados = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({'status': 'erro', 'mensagem': f"Erro ao parsear o JSON: {e}"}, status=400)

        data_inicio = timezone.datetime(2024, 9, 19)  # Data de início da verificação
        data_atual = timezone.now().date()  # Data atual para verificar até quando gerar os resumos

        # Filtra os usuários que estão no grupo "planificador" e exclui os usuários com id 3, 9 e 10
        usuarios = CustomUsuario.objects.filter(groups__name='Planificador').exclude(id__in=[3, 9, 10])

        for usuario in usuarios:
            # Itera por cada dia desde a data de início até a data atual
            data = data_inicio.date()
            while data <= data_atual:
                # Verifica se já existe um resumo para o usuário e a data
                if not ResumoAlteracoes_Planificador.objects.filter(usuario=usuario, data=data).exists():
                    # Busca todas as alterações do usuário naquele dia
                    alteracoes_dia = HistoricoAlteracoes.objects.filter(
                        usuario=usuario, 
                        data_alteracao__date=data
                    )

                    if alteracoes_dia.exists():
                        # Cria um contexto para a IA resumir as alterações, incluindo a escola e o nome amigável dos campos
                        alteracoes_texto = "\n".join(
                            [f"Escola: {alteracao.planificador.escola.nome_da_escola}\n"
                             f"Alteração em {alteracao.data_alteracao.strftime('%H:%M:%S')}:\n"
                             f"{get_verbose_field_names(alteracao.planificador, alteracao.alteracoes)}"
                             for alteracao in alteracoes_dia]
                        )

                        # Gera o resumo usando a API da OpenAI
                        api_key = request.user.api_key
                        resumo = config_resumo_alteracoes(alteracoes_texto, api_key)
                        
                    else:
                        # Se não houver alterações no dia, cria um resumo padrão
                        resumo = "Nenhum ajuste realizado no planificador por este usuário neste dia."

                    # Salva o resumo na model ResumoAlteracoes_Planificador
                    ResumoAlteracoes_Planificador.objects.create(
                        usuario=usuario,
                        data=data,
                        resumo=resumo
                    )

                # Avança para o próximo dia
                data += timedelta(days=1)

        return JsonResponse({'status': 'sucesso', 'mensagem': 'Resumos gerados com sucesso!'}, status=200)

    return JsonResponse({'status': 'falha', 'mensagem': 'Método não permitido'}, status=405)


def gerar_resumo_alteracoes(request):
    data_inicio = timezone.datetime(2024, 9, 19)  # Data de início da verificação
    data_atual = timezone.now().date()  # Data atual para verificar até quando gerar os resumos

    # Filtra os usuários que estão no grupo "planificador" e exclui os usuários com id 3, 9 e 10
    usuarios = CustomUsuario.objects.filter(groups__name='Planificador').exclude(id__in=[3, 9, 10, 21])

    for usuario in usuarios:
        # Itera por cada dia desde a data de início até a data atual
        data = data_inicio.date()
        while data <= data_atual:
            # Verifica se já existe um resumo para o usuário e a data
            if not ResumoAlteracoes_Planificador.objects.filter(usuario=usuario, data=data).exists():
                # Busca todas as alterações do usuário naquele dia
                alteracoes_dia = HistoricoAlteracoes.objects.filter(
                    usuario=usuario, 
                    data_alteracao__date=data
                )

                if alteracoes_dia.exists():
                    # Cria um contexto para a IA resumir as alterações, incluindo a escola e o nome amigável dos campos
                    alteracoes_texto = "\n".join(
                        [f"Escola: {alteracao.planificador.escola.nome_da_escola}\n"
                         f"Alteração em {alteracao.data_alteracao.strftime('%H:%M:%S')}:\n"
                         f"{get_verbose_field_names(alteracao.planificador, alteracao.alteracoes)}"
                         for alteracao in alteracoes_dia]
                    )

                    # Gera o resumo usando a API da OpenAI
                    api_key = request.user.api_key
                    resumo = config_resumo_alteracoes(alteracoes_texto, api_key)
                    
                else:
                    # Se não houver alterações no dia, cria um resumo padrão
                    resumo = "Nenhum ajuste realizado no planificador por este usuário neste dia."

                # Salva o resumo na model ResumoAlteracoes_Planificador
                ResumoAlteracoes_Planificador.objects.create(
                    usuario=usuario,
                    data=data,
                    resumo=resumo
                )

            # Avança para o próximo dia
            data += timedelta(days=1)

    return redirect('controle_escolas')

def get_verbose_field_names(planificador, alteracoes_texto):
    """
    Substitui os nomes das colunas pelos verbose_name dos campos da model.
    """
    campos_verbose = {field.name: field.verbose_name for field in planificador._meta.fields}

    # Substitui os nomes das colunas pelos verbose_name no texto das alterações
    for campo, verbose in campos_verbose.items():
        alteracoes_texto = alteracoes_texto.replace(campo, verbose)

    return alteracoes_texto

######################## BUSCA ESCOLAS PLANIFICADOR #################################################
class EscolaSearchView(ListView):
    model = CRM_FUI
    template_name = "chatapp/planificador/busca_escolas.html"  # Altere para o caminho correto do seu template
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "")
        context["order_by"] = self.request.GET.get("order_by", "nome_da_escola")
        page_obj = context["page_obj"]

        # Obtém o número da página atual
        current_page = page_obj.number

        # Se há mais de 5 páginas
        if page_obj.paginator.num_pages > 5:
            if current_page - 2 < 1:
                start_page = 1
                end_page = 5
            elif current_page + 2 > page_obj.paginator.num_pages:
                start_page = page_obj.paginator.num_pages - 4
                end_page = page_obj.paginator.num_pages
            else:
                start_page = current_page - 2
                end_page = current_page + 2
        else:
            start_page = 1
            end_page = page_obj.paginator.num_pages

        context["page_range"] = range(start_page, end_page + 1)

        # Adiciona porcentagem_planificador ao contexto
        for escola in context['object_list']:
            planificador = Planificador_2024.objects.filter(escola=escola).first()
            if planificador:
                escola.planificador = planificador
                escola.porcentagem_planificador = calcular_porcentagem_sim(planificador)
            else:
                escola.planificador = None
                escola.porcentagem_planificador = 0

        return context

    def get_queryset(self):
        query = self.request.GET.get("q")
        order_by = self.request.GET.get("order_by", "nome_da_escola")
        if query:
            return CRM_FUI.objects.filter(Q(nome_da_escola__icontains=query)).order_by(order_by)
        return CRM_FUI.objects.all().order_by(order_by)
    


############################# PEDIDOS CHINOOK #################################################


@login_required(login_url='/login/')
def import_pedidos(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file or not file.name.endswith(".xlsx"):
            messages.error(request, "Por favor, envie um arquivo Excel.")
            return redirect("import_pedidos")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo Excel: {e}")
            return redirect("import_pedidos")

        with transaction.atomic():
            Pedido.objects.all().delete()  # Limpar a tabela antes de importar novos dados

            for _, row in df.iterrows():
                try:
                    escola = CRM_FUI.objects.get(id_escola=row["id_escola"])
                    
                    Pedido.objects.create(
                        id_linha=row["id_linha"],
                        ano=row["ano"],
                        escola=escola,
                        data_do_pedido=row["data_do_pedido"],
                        serie=row["serie"],
                        nome_do_produto=row["nome_do_produto"],
                        codigo_do_produto=row["codigo_do_produto"],
                        nome_pais=row["nome_pais"],
                        nome_do_aluno=row["nome_do_aluno"],
                        pedido=row["pedido"],
                        status_erp=row["status_erp"],
                        status_logistico=row["status_logistico"],
                        status_transportadora=row["status_transportadora"],
                    )
                except CRM_FUI.DoesNotExist:
                    messages.error(
                        request, f"Escola com id_escola {row['id_escola']} não encontrada."
                    )
                    continue
                except Exception as e:
                    messages.error(
                        request,
                        f"Erro ao importar a linha com id_linha {row['id_linha']}: {e}",
                    )
                    continue

        messages.success(request, "Dados importados com sucesso!")
        return redirect("import_pedidos_chinook")
    return render(request, "chatapp/import/import_pedidos.html")

################################################# ATUALIZAÇÃO JSON ######################################################


@csrf_exempt
def import_vendas_slm_2025_json(request):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({'status': 'erro', 'mensagem': f"Erro ao parsear o JSON: {e}"}, status=400)

        with transaction.atomic():
            for result in dados.get('results', []):
                for table in result.get('tables', []):
                    for row in table.get('rows', []):
                        try:
                            escola = CRM_FUI.objects.get(id_escola=row.get('fui_slm_2025[id_escola]'))

                            # Converte o campo data_do_pedido para um objeto datetime
                            data_do_pedido = datetime.strptime(row.get('fui_slm_2025[data_do_pedido]', ''), "%Y-%m-%dT%H:%M:%S")

                            # Cria o novo registro com o novo campo id_linha
                            Vendas_SLM_2025.objects.create(
                                escola=escola,
                                numero_do_pedido=row.get('fui_slm_2025[numero_do_pedido]', ''),
                                nome_pais=row.get('fui_slm_2025[nome_pais]', ''),
                                nome_do_aluno=row.get('fui_slm_2025[nome_do_aluno]', ''),
                                data_do_pedido=data_do_pedido,
                                quantidade=row.get('[Sumquantidade]', 0),
                                id_linha=row.get('fui_slm_2025[id_linha]', 0)  # Novo campo adicionado
                            )
                        except CRM_FUI.DoesNotExist:
                            return JsonResponse(
                                {'status': 'erro', 'mensagem': f"Escola com id_escola {row.get('fui_slm_2025[id_escola]')} não encontrada."},
                                status=400
                            )
                        except Exception as e:
                            return JsonResponse(
                                {'status': 'erro', 'mensagem': f"Erro ao importar o pedido {row.get('fui_slm_2025[numero_do_pedido]')}: {e}"},
                                status=500
                            )

        return JsonResponse({'status': 'sucesso', 'mensagem': 'Dados importados com sucesso!'}, status=200)

    return JsonResponse({'status': 'falha', 'mensagem': 'Método não permitido'}, status=405)


@csrf_exempt
def delete_vendas_slm_2025_json(request):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({'status': 'erro', 'mensagem': f"Erro ao parsear o JSON: {e}"}, status=400)

        with transaction.atomic():
            for result in dados.get('results', []):
                for table in result.get('tables', []):
                    for row in table.get('rows', []):
                        try:
                            # Pega o ID da venda
                            id_venda = row.get('public ia_vendas_slm_2025 - retirar[id]', None)
                            if id_venda is not None:
                                # Tenta buscar e deletar o registro correspondente
                                Vendas_SLM_2025.objects.filter(id=id_venda).delete()
                            else:
                                return JsonResponse(
                                    {'status': 'erro', 'mensagem': 'ID da venda não encontrado no JSON.'},
                                    status=400
                                )
                        except Exception as e:
                            return JsonResponse(
                                {'status': 'erro', 'mensagem': f"Erro ao tentar excluir a venda com ID {id_venda}: {e}"},
                                status=500
                            )

        return JsonResponse({'status': 'sucesso', 'mensagem': 'Vendas excluídas com sucesso!'}, status=200)

    return JsonResponse({'status': 'falha', 'mensagem': 'Método não permitido'}, status=405)

@csrf_exempt
def import_vendas_slm_2024_json(request):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({'status': 'erro', 'mensagem': f"Erro ao parsear o JSON: {e}"}, status=400)

        with transaction.atomic():
            for result in dados.get('results', []):
                for table in result.get('tables', []):
                    for row in table.get('rows', []):
                        try:
                            escola = CRM_FUI.objects.get(id_escola=row.get('slm_2024[id_escola]'))
                            data_do_pedido = datetime.strptime(row.get('slm_2024[data_do_pedido]', ''), "%Y-%m-%dT%H:%M:%S")

                            # Atualiza ou cria a venda com base no numero_do_pedido e escola
                            Vendas_SLM_2024.objects.update_or_create(
                                numero_do_pedido=row.get('slm_2024[numero_do_pedido]', ''),
                                escola=escola,
                                defaults={
                                    "nome_pais": row.get('slm_2024[nome_pais]', ''),
                                    "nome_do_aluno": row.get('slm_2024[nome_do_aluno]', ''),
                                    "data_do_pedido": data_do_pedido,
                                    "quantidade": row.get('[Sumquantidade]', 0),
                                    "id_linha": row.get('slm_2024[id_linha]', None)
                                }
                            )
                        except CRM_FUI.DoesNotExist:
                            return JsonResponse(
                                {'status': 'erro', 'mensagem': f"Escola com id_escola {row.get('slm_2024[id_escola]')} não encontrada."},
                                status=400
                            )
                        except Exception as e:
                            return JsonResponse(
                                {'status': 'erro', 'mensagem': f"Erro ao importar o pedido {row.get('slm_2024[numero_do_pedido]')}: {e}"},
                                status=500
                            )

        return JsonResponse({'status': 'sucesso', 'mensagem': 'Dados importados e atualizados com sucesso!'}, status=200)

    return JsonResponse({'status': 'falha', 'mensagem': 'Método não permitido'}, status=405)

@csrf_exempt
def delete_vendas_slm_2024_json(request):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({'status': 'erro', 'mensagem': f"Erro ao parsear o JSON: {e}"}, status=400)

        with transaction.atomic():
            for result in dados.get('results', []):
                for table in result.get('tables', []):
                    for row in table.get('rows', []):
                        try:
                            # Pega o ID da venda
                            id_venda = row.get('public ia_vendas_slm_2024 - retirar[id]', None)
                            if id_venda is not None:
                                # Tenta buscar e deletar o registro correspondente
                                Vendas_SLM_2024.objects.filter(id=id_venda).delete()
                            else:
                                return JsonResponse(
                                    {'status': 'erro', 'mensagem': 'ID da venda não encontrado no JSON.'},
                                    status=400
                                )
                        except Exception as e:
                            return JsonResponse(
                                {'status': 'erro', 'mensagem': f"Erro ao tentar excluir a venda com ID {id_venda}: {e}"},
                                status=500
                            )

        return JsonResponse({'status': 'sucesso', 'mensagem': 'Vendas excluídas com sucesso!'}, status=200)

    return JsonResponse({'status': 'falha', 'mensagem': 'Método não permitido'}, status=405)



@csrf_exempt
def import_crm_fui_json(request):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({'status': 'erro', 'mensagem': f"Erro ao parsear o JSON: {e}"}, status=400)

        with transaction.atomic():
            for result in dados.get('results', []):
                for table in result.get('tables', []):
                    for row in table.get('rows', []):
                        try:
                            CRM_FUI.objects.update_or_create(
                                id_escola=row.get('CRM_B2B[id_escola]'),
                                defaults={
                                    "nome_da_escola": row.get('CRM_B2B[nome_da_escola]', ''),
                                    "CNPJ": row.get('CRM_B2B[CNPJ]', ''),
                                    "status_da_escola": row.get('CRM_B2B[status_da_escola]', ''),
                                    "slms_vendidos": row.get('CRM_B2B[slms_vendidos]', 0),
                                    "slms_vendidos_25": row.get('CRM_B2B[slms_vendidos_25]', 0),
                                    "meta": row.get('CRM_B2B[meta]', 0),
                                    "cluster": row.get('CRM_B2B[cluster]', ''),
                                    "cep_escola": row.get('CRM_B2B[cep_escola]', ''),
                                    "endereco": row.get('CRM_B2B[endereco]', ''),
                                    "complemento_escola": row.get('CRM_B2B[complemento_escola]', ''),
                                    "bairro_escola": row.get('CRM_B2B[bairro_escola]', ''),
                                    "cidade_da_escola": row.get('CRM_B2B[cidade_da_escola]', ''),
                                    "estado_da_escola": row.get('CRM_B2B[estado_da_escola]', ''),
                                    "regiao_da_escola": row.get('CRM_B2B[regiao_da_escola]', ''),
                                    "status_de_adimplencia": row.get('CRM_B2B[status_de_adimplencia]', ''),
                                    "inadimplencia": row.get('CRM_B2B[inadimplencia]', 0),
                                    "ticket_medio": row.get('CRM_B2B[ticket_medio]', 0),
                                    "valor_royalties": row.get('CRM_B2B[valor_royalties]', 0),
                                    "valor_fdmp": row.get('CRM_B2B[valor_fdmp]', 0),
                                    "segmento_da_escola": row.get('CRM_B2B[segmento_da_escola]', ''),
                                    "atual_serie": row.get('CRM_B2B[atual_serie]', ''),
                                    "avanco_segmento": row.get('CRM_B2B[avanco_segmento]', ''),
                                    "telefone_de_contato_da_escola": row.get('CRM_B2B[telefone_de_contato_da_escola]', ''),
                                    "email_da_escola": row.get('CRM_B2B[email_da_escola]', ''),
                                    "nps_pais_2024_1_onda": row.get('CRM_B2B[nps_pais_2024_1_onda]', 0),
                                    "quality_assurance_2024": row.get('CRM_B2B[quality_assurance_2024]', 0),
                                    "cliente_oculto_2024": row.get('CRM_B2B[cliente_oculto_2024]', 0),
                                    "consultor_saf": row.get('CRM_B2B[consultor_saf]', ''),
                                    "consultor_academico": row.get('CRM_B2B[consultor_academico]', ''),
                                    "consultor_comercial": row.get('CRM_B2B[consultor_comercial]', ''),
                                    "consultor_gestao_escolar": row.get('CRM_B2B[consultor_gestao_escolar]', ''),
                                    "dias_uteis_entrega_slm": row.get('CRM_B2B[dias_uteis_entrega_slm]', ''),
                                }
                            )
                        except Exception as e:
                            return JsonResponse(
                                {'status': 'erro', 'mensagem': f"Erro ao importar o registro com id_escola {row.get('CRM_B2B[id_escola]')}: {e}"},
                                status=500
                            )

        return JsonResponse({'status': 'sucesso', 'mensagem': 'Dados importados com sucesso!'}, status=200)

    return JsonResponse({'status': 'falha', 'mensagem': 'Método não permitido'}, status=405)




@csrf_exempt
def import_pedidos_alterados_json(request):
    if request.method == 'POST':
        try:
            dados = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({'status': 'erro', 'mensagem': f"Erro ao parsear o JSON: {e}"}, status=400)

        # Exclui todos os registros existentes
        PedidosAlterados.objects.all().delete()

        with transaction.atomic():
            for result in dados.get('results', []):
                for table in result.get('tables', []):
                    for row in table.get('rows', []):
                        try:
                            # Obtém a escola com base no id_escola
                            escola = CRM_FUI.objects.get(id_escola=row.get('CANCELADOS PARCIALMENTE[id_escola]'))

                            # Cria um novo registro vinculado à escola
                            PedidosAlterados.objects.create(
                                escola=escola,
                                nome_do_aluno=row.get('CANCELADOS PARCIALMENTE[nome_do_aluno]', ''),
                                numero_do_pedido=row.get('CANCELADOS PARCIALMENTE[numero_do_pedido]', ''),
                                motivo=row.get('CANCELADOS PARCIALMENTE[motivo]', ''),
                                alterado_por=row.get('CANCELADOS PARCIALMENTE[alterado_por]', '')
                            )
                        except CRM_FUI.DoesNotExist:
                            return JsonResponse(
                                {'status': 'erro', 'mensagem': f"Escola com id_escola {row.get('CANCELADOS PARCIALMENTE[id_escola]')} não encontrada."},
                                status=400
                            )
                        except Exception as e:
                            return JsonResponse(
                                {'status': 'erro', 'mensagem': f"Erro ao importar o pedido {row.get('CANCELADOS PARCIALMENTE[numero_do_pedido]')}: {e}"},
                                status=500
                            )

        return JsonResponse({'status': 'sucesso', 'mensagem': 'Dados importados com sucesso!'}, status=200)

    return JsonResponse({'status': 'falha', 'mensagem': 'Método não permitido'}, status=405)




##################################################################################################
class ReclamacaoCreateView(LoginRequiredMixin, View):
    login_url = '/login/'

    def get(self, request):
        form = ReclamacaoForm()
        return render(request, "chatapp/reclamacao/reclamacao_form.html", {"form": form})

    def post(self, request):
        form = ReclamacaoForm(request.POST)
        if form.is_valid():
            reclamacao = form.save(commit=False)
            if reclamacao.status == 'finalizado':
                reclamacao.data_conclusao = datetime.now().date()  # Preenche com a data atual
            reclamacao.save()
            messages.success(request, "Reclamacao criada com sucesso!")
            return redirect("search_escolas_reclamacao")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erro no campo {form.fields[field].label}: {error}")
        return render(request, "chatapp/reclamacao/reclamacao_form.html", {"form": form})

# View de edição
class ReclamacaoUpdateView(LoginRequiredMixin, View):
    login_url = '/login/'

    def get(self, request, pk):
        reclamacao = get_object_or_404(Reclamacao, pk=pk)
        form = ReclamacaoForm(instance=reclamacao)
        return render(request, "chatapp/reclamacao/reclamacao_form.html", {"form": form})

    def post(self, request, pk):
        reclamacao = get_object_or_404(Reclamacao, pk=pk)
        form = ReclamacaoForm(request.POST, instance=reclamacao)
        if form.is_valid():
            reclamacao = form.save(commit=False)
            if reclamacao.status == 'finalizado':
                reclamacao.data_conclusao = datetime.now().date()  # Preenche com a data atual
            reclamacao.save()
            messages.success(request, "Reclamacao atualizada com sucesso!")
            return redirect("search_escolas_reclamacao")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erro no campo {form.fields[field].label}: {error}")
        return render(request, "chatapp/reclamacao/reclamacao_form.html", {"form": form})


class Escola_Reclamacao_SearchView(ListView):
    model = Reclamacao
    template_name = "chatapp/reclamacao/busca_escolas_reclamacao.html"
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "")
        context["order_by"] = self.request.GET.get("order_by", "escola__nome_da_escola")
        page_obj = context["page_obj"]

        # Lógica de paginação personalizada
        current_page = page_obj.number
        if page_obj.paginator.num_pages > 5:
            if current_page - 2 < 1:
                start_page = 1
                end_page = 5
            elif current_page + 2 > page_obj.paginator.num_pages:
                start_page = page_obj.paginator.num_pages - 4
                end_page = page_obj.paginator.num_pages
            else:
                start_page = current_page - 2
                end_page = current_page + 2
        else:
            start_page = 1
            end_page = page_obj.paginator.num_pages

        context["page_range"] = range(start_page, end_page + 1)
        return context

    def get_queryset(self):
        query = self.request.GET.get("q")
        order_by = self.request.GET.get("order_by", "escola__nome_da_escola")
        queryset = Reclamacao.objects.all()

        if query:
            # Filtra pelos campos que você mencionou e pelo status
            if query.lower() == "concluido":
                queryset = queryset.filter(status="finalizado")
            elif query.lower() == "pendente":
                queryset = queryset.exclude(status="finalizado")
            else:
                queryset = queryset.filter(
                    Q(escola__id_escola__icontains=query) |
                    Q(escola__nome_da_escola__icontains=query) |
                    Q(escola__cluster__icontains=query) |  # Busca pelo Cluster
                    Q(titulo__icontains=query)
                )

        return queryset.order_by(order_by)
#################################################################################################


def atualizar_id_escola_view(request):
    if request.method == 'POST':
        form = AtualizarIDEscolaForm(request.POST)
        if form.is_valid():
            id_escola_atual = form.cleaned_data['id_escola_atual'].id_escola
            novo_id_escola = form.cleaned_data['novo_id_escola']
            try:
                with transaction.atomic():
                    # Atualiza o ID na tabela CRM_FUI
                    crm_fui = CRM_FUI.objects.get(id_escola=id_escola_atual)
                    crm_fui.id_escola = novo_id_escola
                    crm_fui.save()

                    # Atualiza referências em outras tabelas
                    Visita_Escola.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    Resumo_Visita_Escola.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    Respostas_NPS.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    Resumo_Respostas_NPS.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    Avaliacao_Cliente_Oculto_24.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    Resumo_Respostas_ClienteOculto24.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    Vendas_SLM_2024.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    Vendas_SLM_2025.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    PedidosAlterados.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    Ticket_Sprinklr.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)
                    Planificador_2024.objects.filter(escola_id=id_escola_atual).update(escola_id=novo_id_escola)

                    messages.success(request, f"ID da escola atualizado de {id_escola_atual} para {novo_id_escola} com sucesso!")
                    return redirect('atualizar_id_escola')  # Redireciona para a mesma página após o sucesso
            except Exception as e:
                messages.error(request, f"Erro ao atualizar o ID: {str(e)}")
    else:
        form = AtualizarIDEscolaForm()

    return render(request, 'atualizar_id_escola.html', {'form': form})



def gerar_resumos_base_conhecimento(request):
    """
    Gera resumos detalhados para todas as entradas da model Base_de_Conhecimento_Geral,
    garantindo que todas as informações do conteúdo original estejam presentes no resumo.
    """
    entradas = Base_de_Conhecimento_Geral.objects.all()
    resultados = []

    for entrada in entradas:
        try:
            print(f"Gerando resumo para a entrada ID: {entrada.id} - Título: {entrada.titulo}")

            # Gera o resumo com todas as informações presentes no conteúdo
            conteudo = entrada.conteudo
            api_key = request.user.api_key  # Certifique-se de que o usuário tem uma chave API válida
            resumo = gerar_resumo_base_de_conhecimento(conteudo, api_key)

            # Salva o resumo na model
            entrada.resumo = resumo
            entrada.save()

            resultados.append(f"Resumo gerado para ID: {entrada.id} - Título: {entrada.titulo}")
            print(f"Resumo gerado com sucesso para ID: {entrada.id}")
        
        except Exception as e:
            print(f"Erro ao gerar resumo para ID: {entrada.id} - Título: {entrada.titulo}: {str(e)}")
            resultados.append(f"Erro ao gerar resumo para ID: {entrada.id} - Título: {entrada.titulo}")

        # Delay opcional para evitar sobrecarga na API
        time.sleep(2)

    print("Processo concluído para todas as entradas.")
    return JsonResponse({"resultados": resultados})